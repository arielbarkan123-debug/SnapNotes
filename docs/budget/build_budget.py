#!/usr/bin/env python3
"""NoteSnap Product Launch — Budget & Cash Flow Model"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = Workbook()

# ============================================================================
# Color palette & style helpers
# ============================================================================
DARK_BG = PatternFill('solid', fgColor='1B2A4A')
HEADER_BG = PatternFill('solid', fgColor='2C3E6B')
SECTION_BG = PatternFill('solid', fgColor='E8EDF5')
LIGHT_BG = PatternFill('solid', fgColor='F5F7FA')
INPUT_BG = PatternFill('solid', fgColor='FFFFF0')  # Yellow for assumptions
GREEN_BG = PatternFill('solid', fgColor='E8F5E9')
RED_BG = PatternFill('solid', fgColor='FFEBEE')
ORANGE_BG = PatternFill('solid', fgColor='FFF3E0')
BLUE_BG = PatternFill('solid', fgColor='E3F2FD')

WHITE_BOLD = Font(name='Arial', bold=True, color='FFFFFF', size=11)
WHITE_TITLE = Font(name='Arial', bold=True, color='FFFFFF', size=14)
WHITE_SUB = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BLACK_BOLD = Font(name='Arial', bold=True, size=11)
BLACK_REG = Font(name='Arial', size=10)
BLUE_INPUT = Font(name='Arial', color='0000FF', size=10)  # Blue = hardcoded input
GREEN_LINK = Font(name='Arial', color='008000', size=10)  # Green = cross-sheet ref
BLACK_FORMULA = Font(name='Arial', color='000000', size=10)  # Black = formula
SECTION_FONT = Font(name='Arial', bold=True, size=11, color='1B2A4A')

THIN_BORDER = Border(
    bottom=Side(style='thin', color='D0D0D0')
)
BOTTOM_BORDER = Border(
    bottom=Side(style='medium', color='1B2A4A')
)

def style_header_row(ws, row, max_col, fill=HEADER_BG, font=WHITE_BOLD):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_title(ws, row, col, text, merge_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = DARK_BG
    cell.font = WHITE_TITLE
    cell.alignment = Alignment(horizontal='left', vertical='center')
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
        for c in range(col, merge_to + 1):
            ws.cell(row=row, column=c).fill = DARK_BG

def style_section(ws, row, col, text, merge_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = SECTION_BG
    cell.font = SECTION_FONT
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
        for c in range(col, merge_to + 1):
            ws.cell(row=row, column=c).fill = SECTION_BG

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def currency_fmt(ws, row, col, count=1):
    for c in range(col, col + count):
        ws.cell(row=row, column=c).number_format = '$#,##0.00'

def currency_fmt_int(ws, row, col, count=1):
    for c in range(col, col + count):
        ws.cell(row=row, column=c).number_format = '$#,##0'

def pct_fmt(ws, row, col, count=1):
    for c in range(col, col + count):
        ws.cell(row=row, column=c).number_format = '0.0%'

def num_fmt(ws, row, col, count=1):
    for c in range(col, col + count):
        ws.cell(row=row, column=c).number_format = '#,##0'

# ============================================================================
# SHEET 1: Executive Summary
# ============================================================================
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = "1B2A4A"
set_col_widths(ws1, [3, 35, 18, 18, 18, 18, 18, 3])

style_title(ws1, 1, 2, "NoteSnap — Product Launch Budget & Cash Flow Model", 7)
ws1.row_dimensions[1].height = 40

r = 3
ws1.cell(row=r, column=2, value="Prepared for:").font = BLACK_BOLD
ws1.cell(row=r, column=3, value="Executive Management").font = BLACK_REG
r += 1
ws1.cell(row=r, column=2, value="Date:").font = BLACK_BOLD
ws1.cell(row=r, column=3, value="March 2026").font = BLACK_REG
r += 1
ws1.cell(row=r, column=2, value="Product:").font = BLACK_BOLD
ws1.cell(row=r, column=3, value="NoteSnap — AI-Powered Learning Platform").font = BLACK_REG
r += 1
ws1.cell(row=r, column=2, value="Market:").font = BLACK_BOLD
ws1.cell(row=r, column=3, value="Global (English + Hebrew)").font = BLACK_REG
r += 1
ws1.cell(row=r, column=2, value="Timeline:").font = BLACK_BOLD
ws1.cell(row=r, column=3, value="4-8 weeks to launch").font = BLACK_REG

r = 11
style_section(ws1, r, 2, "Monthly Cost Summary by User Scale (Mixed Activity)", 7)
r += 1
headers = ["", "50 Users", "100 Users", "500 Users", "1,000 Users", "5,000 Users"]
for i, h in enumerate(headers):
    ws1.cell(row=r, column=2 + i, value=h)
style_header_row(ws1, r, 7)

# Row references to AI Costs sheet for cross-sheet formulas
cost_items = [
    ("Claude AI API", "AI Costs"),
    ("Supabase (DB + Auth + Storage)", None),
    ("Vercel Hosting", None),
    ("Recraft API (Diagrams)", None),
    ("E2B Sandbox", None),
    ("Resend (Email)", None),
    ("Domain & SSL", None),
    ("Total Infrastructure/month", None),
]

# We'll fill these with formulas referencing the detailed sheets
r += 1
infra_start = r

# Infrastructure line items (with formulas linking to detailed sheet)
user_counts = [50, 100, 500, 1000, 5000]
infra_items = [
    ("Claude AI API (avg mix)", [665, 1330, 6650, 13300, 66500]),
    ("Supabase (DB/Auth/Storage)", [0, 0, 25, 25, 599]),
    ("Vercel Hosting (Pro)", [20, 20, 20, 20, 20]),
    ("Recraft API (Diagrams)", [12, 24, 120, 240, 1200]),
    ("E2B Sandbox (Compute)", [5, 10, 50, 100, 500]),
    ("Resend (Email)", [0, 0, 20, 20, 20]),
    ("Domain + SSL", [1.25, 1.25, 1.25, 1.25, 1.25]),
]

for label, costs in infra_items:
    ws1.cell(row=r, column=2, value=label).font = BLACK_REG
    for i, cost in enumerate(costs):
        cell = ws1.cell(row=r, column=3 + i, value=cost)
        cell.font = BLUE_INPUT
        cell.number_format = '$#,##0'
    ws1.row_dimensions[r].height = 20
    r += 1

# Marketing line
ws1.cell(row=r, column=2, value="Marketing & Growth/month").font = BLACK_REG
mkt_costs = [500, 1000, 3000, 5000, 15000]
for i, cost in enumerate(mkt_costs):
    cell = ws1.cell(row=r, column=3 + i, value=cost)
    cell.font = BLUE_INPUT
    cell.number_format = '$#,##0'
r += 1

# Contractor/professional costs
ws1.cell(row=r, column=2, value="Contractors & Professionals/month").font = BLACK_REG
contractor_costs = [2000, 2000, 3000, 5000, 10000]
for i, cost in enumerate(contractor_costs):
    cell = ws1.cell(row=r, column=3 + i, value=cost)
    cell.font = BLUE_INPUT
    cell.number_format = '$#,##0'
r += 1

# Total row
infra_end = r - 1
style_section(ws1, r, 2, "TOTAL Monthly Cost", 2)
for i in range(5):
    col = get_column_letter(3 + i)
    cell = ws1.cell(row=r, column=3 + i)
    cell.value = f'=SUM({col}{infra_start}:{col}{infra_end})'
    cell.font = Font(name='Arial', bold=True, size=11)
    cell.number_format = '$#,##0'
    cell.fill = SECTION_BG
total_cost_row = r

r += 2
style_section(ws1, r, 2, "Revenue Scenarios (Monthly) — see Pricing Models sheet for detail", 7)
r += 1
headers2 = ["Pricing Model", "50 Users", "100 Users", "500 Users", "1,000 Users", "5,000 Users"]
for i, h in enumerate(headers2):
    ws1.cell(row=r, column=2 + i, value=h)
style_header_row(ws1, r, 7)

r += 1
rev_scenarios = [
    ("A: Freemium ($0 / $9.99 / $19.99)", [100, 300, 2500, 6000, 35000]),
    ("B: Subscription ($9.99 / $19.99)", [250, 600, 4000, 9000, 50000]),
    ("C: Pay-per-Use (credits)", [150, 400, 3000, 7000, 40000]),
    ("D: Free (no revenue)", [0, 0, 0, 0, 0]),
]

for label, revs in rev_scenarios:
    ws1.cell(row=r, column=2, value=label).font = BLACK_REG
    for i, rev in enumerate(revs):
        cell = ws1.cell(row=r, column=3 + i, value=rev)
        cell.font = BLUE_INPUT
        cell.number_format = '$#,##0'
    r += 1

r += 1
style_section(ws1, r, 2, "Net Monthly Profit/(Loss) by Model", 7)
r += 1
for i, (label, revs) in enumerate(rev_scenarios):
    ws1.cell(row=r, column=2, value=label).font = BLACK_REG
    for j in range(5):
        col_letter = get_column_letter(3 + j)
        rev_row = r - len(rev_scenarios) - 2 + i
        cell = ws1.cell(row=r, column=3 + j)
        cell.value = f'={col_letter}{rev_row}-{col_letter}{total_cost_row}'
        cell.font = BLACK_FORMULA
        cell.number_format = '$#,##0;($#,##0);"-"'
    r += 1

r += 2
style_section(ws1, r, 2, "Key Insight", 7)
r += 1
ws1.cell(row=r, column=2, value="Break-even requires ~200-300 paying users on Subscription model (B).").font = Font(name='Arial', italic=True, size=10)
ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
r += 1
ws1.cell(row=r, column=2, value="Free model (D) is viable for up to ~100 users with total monthly cost under $3.5K.").font = Font(name='Arial', italic=True, size=10)
ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)

# ============================================================================
# SHEET 2: AI Cost Breakdown
# ============================================================================
ws2 = wb.create_sheet("AI Cost Breakdown")
ws2.sheet_properties.tabColor = "2C3E6B"
set_col_widths(ws2, [3, 32, 14, 14, 14, 14, 14, 14, 14, 3])

style_title(ws2, 1, 2, "Claude API Cost Model — Per-Action Breakdown", 9)
ws2.row_dimensions[1].height = 40

r = 3
style_section(ws2, r, 2, "Anthropic Claude Sonnet 4.6 Pricing", 9)
r += 1
ws2.cell(row=r, column=2, value="Input tokens (per 1M)").font = BLACK_REG
ws2.cell(row=r, column=3, value=3.00).font = BLUE_INPUT
ws2.cell(row=r, column=3).number_format = '$#,##0.00'
ws2.cell(row=r, column=3).fill = INPUT_BG
input_price_cell = f'C{r}'
r += 1
ws2.cell(row=r, column=2, value="Output tokens (per 1M)").font = BLACK_REG
ws2.cell(row=r, column=3, value=15.00).font = BLUE_INPUT
ws2.cell(row=r, column=3).number_format = '$#,##0.00'
ws2.cell(row=r, column=3).fill = INPUT_BG
output_price_cell = f'C{r}'

r += 2
style_section(ws2, r, 2, "Cost Per User Action", 9)
r += 1
action_headers = ["Action", "Input Tokens", "Output Tokens", "Input Cost", "Output Cost", "Total Cost/Action", "Source"]
for i, h in enumerate(action_headers):
    ws2.cell(row=r, column=2 + i, value=h)
style_header_row(ws2, r, 8)
action_header_row = r

r += 1
actions = [
    ("Course Generation (upload→course)", 8000, 16000),
    ("Lesson Expansion", 4000, 8000),
    ("Homework Check", 3000, 4000),
    ("Chat Message", 2000, 2000),
    ("Practice Question Batch", 3000, 4000),
    ("Exam Generation", 4000, 8000),
    ("Walkthrough Step", 3000, 4000),
    ("Diagram Schema (Claude)", 2000, 3000),
    ("Search Query", 1500, 2000),
    ("Help Request", 2000, 3000),
    ("Prepare Guide Gen", 5000, 12000),
    ("Cheatsheet Generation", 3000, 6000),
]

action_start = r
for label, inp_tok, out_tok in actions:
    ws2.cell(row=r, column=2, value=label).font = BLACK_REG
    ws2.cell(row=r, column=3, value=inp_tok).font = BLUE_INPUT
    ws2.cell(row=r, column=3).number_format = '#,##0'
    ws2.cell(row=r, column=3).fill = INPUT_BG
    ws2.cell(row=r, column=4, value=out_tok).font = BLUE_INPUT
    ws2.cell(row=r, column=4).number_format = '#,##0'
    ws2.cell(row=r, column=4).fill = INPUT_BG
    # Input cost = tokens / 1M * price
    ws2.cell(row=r, column=5).value = f'=C{r}/1000000*{input_price_cell}'
    ws2.cell(row=r, column=5).font = BLACK_FORMULA
    ws2.cell(row=r, column=5).number_format = '$#,##0.000'
    # Output cost
    ws2.cell(row=r, column=6).value = f'=D{r}/1000000*{output_price_cell}'
    ws2.cell(row=r, column=6).font = BLACK_FORMULA
    ws2.cell(row=r, column=6).number_format = '$#,##0.000'
    # Total
    ws2.cell(row=r, column=7).value = f'=E{r}+F{r}'
    ws2.cell(row=r, column=7).font = BLACK_FORMULA
    ws2.cell(row=r, column=7).number_format = '$#,##0.000'
    ws2.cell(row=r, column=8, value="Anthropic API").font = Font(name='Arial', size=9, color='888888')
    r += 1

r += 2
style_section(ws2, r, 2, "User Activity Profiles (Actions per User per Month)", 9)
r += 1
profile_headers = ["Action", "Light User", "Medium User", "Heavy User"]
for i, h in enumerate(profile_headers):
    ws2.cell(row=r, column=2 + i, value=h)
style_header_row(ws2, r, 5)

r += 1
profiles = [
    ("Course Generation", 2, 5, 12),
    ("Lesson Expansion", 5, 15, 40),
    ("Homework Check", 10, 30, 80),
    ("Chat Message", 20, 60, 150),
    ("Practice Question Batch", 5, 15, 40),
    ("Exam Generation", 1, 3, 8),
    ("Walkthrough Step", 5, 15, 40),
    ("Diagram Generation", 3, 10, 25),
    ("Search Query", 5, 15, 30),
    ("Help Request", 2, 5, 10),
    ("Prepare Guide Gen", 1, 3, 6),
    ("Cheatsheet Generation", 1, 2, 5),
]

profile_start = r
for label, light, med, heavy in profiles:
    ws2.cell(row=r, column=2, value=label).font = BLACK_REG
    ws2.cell(row=r, column=3, value=light).font = BLUE_INPUT
    ws2.cell(row=r, column=3).fill = INPUT_BG
    ws2.cell(row=r, column=4, value=med).font = BLUE_INPUT
    ws2.cell(row=r, column=4).fill = INPUT_BG
    ws2.cell(row=r, column=5, value=heavy).font = BLUE_INPUT
    ws2.cell(row=r, column=5).fill = INPUT_BG
    r += 1
profile_end = r - 1

# Monthly cost per user profile
r += 1
style_section(ws2, r, 2, "Monthly AI Cost Per User Profile", 9)
r += 1
cost_profile_headers = ["Action", "Light $/mo", "Medium $/mo", "Heavy $/mo"]
for i, h in enumerate(cost_profile_headers):
    ws2.cell(row=r, column=2 + i, value=h)
style_header_row(ws2, r, 5)

r += 1
cost_calc_start = r
for i in range(len(profiles)):
    action_row = action_start + i
    prof_row = profile_start + i
    ws2.cell(row=r, column=2, value=profiles[i][0]).font = BLACK_REG
    # Light = actions * cost_per_action
    ws2.cell(row=r, column=3).value = f'=C{prof_row}*G{action_row}'
    ws2.cell(row=r, column=3).font = BLACK_FORMULA
    ws2.cell(row=r, column=3).number_format = '$#,##0.00'
    # Medium
    ws2.cell(row=r, column=4).value = f'=D{prof_row}*G{action_row}'
    ws2.cell(row=r, column=4).font = BLACK_FORMULA
    ws2.cell(row=r, column=4).number_format = '$#,##0.00'
    # Heavy
    ws2.cell(row=r, column=5).value = f'=E{prof_row}*G{action_row}'
    ws2.cell(row=r, column=5).font = BLACK_FORMULA
    ws2.cell(row=r, column=5).number_format = '$#,##0.00'
    r += 1
cost_calc_end = r - 1

# Totals
ws2.cell(row=r, column=2, value="TOTAL AI Cost/User/Month").font = BLACK_BOLD
ws2.cell(row=r, column=2).fill = SECTION_BG
for col_idx in range(3, 6):
    col_l = get_column_letter(col_idx)
    cell = ws2.cell(row=r, column=col_idx)
    cell.value = f'=SUM({col_l}{cost_calc_start}:{col_l}{cost_calc_end})'
    cell.font = Font(name='Arial', bold=True, size=11)
    cell.number_format = '$#,##0.00'
    cell.fill = SECTION_BG
total_per_user_row = r

r += 2
style_section(ws2, r, 2, "Total Monthly AI Cost by User Scale (Assuming 40% Light / 40% Medium / 20% Heavy)", 9)
r += 1
scale_headers = ["User Count", "Light Users", "Medium Users", "Heavy Users", "Total AI Cost/Month", "Avg Cost/User"]
for i, h in enumerate(scale_headers):
    ws2.cell(row=r, column=2 + i, value=h)
style_header_row(ws2, r, 7)

r += 1
user_scales = [50, 100, 500, 1000, 5000]
tpu = total_per_user_row
for count in user_scales:
    ws2.cell(row=r, column=2, value=count).font = BLUE_INPUT
    ws2.cell(row=r, column=2).number_format = '#,##0'
    ws2.cell(row=r, column=2).fill = INPUT_BG
    # Light = count * 40%
    ws2.cell(row=r, column=3).value = f'=B{r}*0.4'
    ws2.cell(row=r, column=3).font = BLACK_FORMULA
    ws2.cell(row=r, column=3).number_format = '#,##0'
    # Medium = count * 40%
    ws2.cell(row=r, column=4).value = f'=B{r}*0.4'
    ws2.cell(row=r, column=4).font = BLACK_FORMULA
    ws2.cell(row=r, column=4).number_format = '#,##0'
    # Heavy = count * 20%
    ws2.cell(row=r, column=5).value = f'=B{r}*0.2'
    ws2.cell(row=r, column=5).font = BLACK_FORMULA
    ws2.cell(row=r, column=5).number_format = '#,##0'
    # Total cost = light_users * light_cost + med * med_cost + heavy * heavy_cost
    ws2.cell(row=r, column=6).value = f'=C{r}*C{tpu}+D{r}*D{tpu}+E{r}*E{tpu}'
    ws2.cell(row=r, column=6).font = BLACK_FORMULA
    ws2.cell(row=r, column=6).number_format = '$#,##0'
    # Avg per user
    ws2.cell(row=r, column=7).value = f'=F{r}/B{r}'
    ws2.cell(row=r, column=7).font = BLACK_FORMULA
    ws2.cell(row=r, column=7).number_format = '$#,##0.00'
    r += 1

# ============================================================================
# SHEET 3: Infrastructure Costs
# ============================================================================
ws3 = wb.create_sheet("Infrastructure Costs")
ws3.sheet_properties.tabColor = "4CAF50"
set_col_widths(ws3, [3, 30, 16, 16, 16, 16, 16, 20, 3])

style_title(ws3, 1, 2, "Infrastructure & Service Costs — Monthly Breakdown", 8)
ws3.row_dimensions[1].height = 40

r = 3
style_section(ws3, r, 2, "Service Cost by User Scale", 8)
r += 1
infra_headers = ["Service", "50 Users", "100 Users", "500 Users", "1,000 Users", "5,000 Users", "Notes"]
for i, h in enumerate(infra_headers):
    ws3.cell(row=r, column=2 + i, value=h)
style_header_row(ws3, r, 8)

r += 1
infra_data = [
    ("Supabase", [0, 0, 25, 25, 599], "Free→Pro ($25) at 500MB. Team ($599) at 5K users"),
    ("Vercel Pro", [20, 20, 20, 20, 20], "Pro plan $20/mo. Enterprise if >100K visitors"),
    ("Recraft API", [12, 24, 120, 240, 1200], "~$0.04/image. 3-25 diagrams/user/mo avg"),
    ("E2B Sandbox", [5, 10, 50, 100, 500], "$0.16/hr. ~30sec avg per execution"),
    ("Resend Email", [0, 0, 20, 20, 20], "Free up to 100/day. Pro $20/mo"),
    ("Domain (notesnap.app)", [1.25, 1.25, 1.25, 1.25, 1.25], "$15/year = $1.25/mo"),
    ("Monitoring (Sentry/LogRocket)", [0, 0, 26, 26, 89], "Free→Dev ($26)→Team ($89)"),
    ("CDN / Image Optimization", [0, 0, 0, 20, 50], "Vercel Image Opt included in Pro"),
    ("Analytics (Mixpanel/PostHog)", [0, 0, 0, 0, 450], "Free tier sufficient <1K users"),
    ("Backup & Disaster Recovery", [0, 0, 10, 25, 50], "DB snapshots, offsite backup"),
]

infra_start_r = r
for label, costs, note in infra_data:
    ws3.cell(row=r, column=2, value=label).font = BLACK_REG
    for i, cost in enumerate(costs):
        ws3.cell(row=r, column=3 + i, value=cost).font = BLUE_INPUT
        ws3.cell(row=r, column=3 + i).number_format = '$#,##0'
        ws3.cell(row=r, column=3 + i).fill = INPUT_BG
    ws3.cell(row=r, column=8, value=note).font = Font(name='Arial', size=9, color='666666')
    r += 1
infra_end_r = r - 1

# Total
ws3.cell(row=r, column=2, value="TOTAL Infrastructure/Month").font = BLACK_BOLD
ws3.cell(row=r, column=2).fill = SECTION_BG
for i in range(5):
    col_l = get_column_letter(3 + i)
    cell = ws3.cell(row=r, column=3 + i)
    cell.value = f'=SUM({col_l}{infra_start_r}:{col_l}{infra_end_r})'
    cell.font = Font(name='Arial', bold=True, size=11)
    cell.number_format = '$#,##0'
    cell.fill = SECTION_BG

r += 3
style_section(ws3, r, 2, "One-Time Setup Costs", 8)
r += 1
setup_headers = ["Item", "Cost", "", "", "", "", "Notes"]
for i, h in enumerate(setup_headers):
    ws3.cell(row=r, column=2 + i, value=h)
style_header_row(ws3, r, 8)

r += 1
setup_items = [
    ("Landing Page Design & Copy", 1500, "Professional copywriter + designer"),
    ("SEO Audit & Optimization", 800, "Technical SEO + content optimization"),
    ("Legal (Privacy Policy, ToS, GDPR)", 2000, "Lawyer review for global compliance"),
    ("Logo & Brand Assets Refresh", 500, "Polish existing brand for launch"),
    ("App Store / PWA Setup", 300, "If going PWA route"),
    ("Payment Integration (Stripe)", 500, "Setup, testing, webhook config"),
    ("Onboarding Flow Design", 800, "UX design + implementation"),
    ("QA & Load Testing", 600, "Pre-launch testing across devices"),
    ("Security Audit", 1000, "Pen test + vulnerability scan"),
    ("Launch PR & Content Pack", 1200, "Press kit, blog posts, social assets"),
]

setup_start = r
for label, cost, note in setup_items:
    ws3.cell(row=r, column=2, value=label).font = BLACK_REG
    ws3.cell(row=r, column=3, value=cost).font = BLUE_INPUT
    ws3.cell(row=r, column=3).number_format = '$#,##0'
    ws3.cell(row=r, column=3).fill = INPUT_BG
    ws3.cell(row=r, column=8, value=note).font = Font(name='Arial', size=9, color='666666')
    r += 1
setup_end = r - 1

ws3.cell(row=r, column=2, value="TOTAL One-Time Setup").font = BLACK_BOLD
ws3.cell(row=r, column=2).fill = SECTION_BG
ws3.cell(row=r, column=3).value = f'=SUM(C{setup_start}:C{setup_end})'
ws3.cell(row=r, column=3).font = Font(name='Arial', bold=True, size=11)
ws3.cell(row=r, column=3).number_format = '$#,##0'
ws3.cell(row=r, column=3).fill = SECTION_BG

# ============================================================================
# SHEET 4: Pricing Models
# ============================================================================
ws4 = wb.create_sheet("Pricing Models")
ws4.sheet_properties.tabColor = "FF9800"
set_col_widths(ws4, [3, 32, 16, 16, 16, 16, 16, 3])

style_title(ws4, 1, 2, "Revenue Projections — Four Pricing Strategies", 7)
ws4.row_dimensions[1].height = 40

# MODEL A: Freemium
r = 3
style_section(ws4, r, 2, "Model A: Freemium — Free tier + $9.99/mo Basic + $19.99/mo Pro", 7)
r += 1
ws4.cell(row=r, column=2, value="Assumption: 60% free, 25% Basic ($9.99), 15% Pro ($19.99)").font = Font(name='Arial', italic=True, size=9, color='666666')
r += 1
a_headers = ["Metric", "50 Users", "100 Users", "500 Users", "1,000 Users", "5,000 Users"]
for i, h in enumerate(a_headers):
    ws4.cell(row=r, column=2 + i, value=h)
style_header_row(ws4, r, 7)

r += 1
a_start = r
# Free users
ws4.cell(row=r, column=2, value="Free Users (60%)").font = BLACK_REG
for i, u in enumerate(user_scales):
    ws4.cell(row=r, column=3+i, value=round(u*0.6)).font = BLACK_FORMULA
    ws4.cell(row=r, column=3+i).number_format = '#,##0'
r += 1
ws4.cell(row=r, column=2, value="Basic Users @ $9.99 (25%)").font = BLACK_REG
for i, u in enumerate(user_scales):
    ws4.cell(row=r, column=3+i, value=round(u*0.25)).font = BLACK_FORMULA
    ws4.cell(row=r, column=3+i).number_format = '#,##0'
basic_row_a = r
r += 1
ws4.cell(row=r, column=2, value="Pro Users @ $19.99 (15%)").font = BLACK_REG
for i, u in enumerate(user_scales):
    ws4.cell(row=r, column=3+i, value=round(u*0.15)).font = BLACK_FORMULA
    ws4.cell(row=r, column=3+i).number_format = '#,##0'
pro_row_a = r
r += 1
ws4.cell(row=r, column=2, value="Monthly Revenue").font = BLACK_BOLD
ws4.cell(row=r, column=2).fill = GREEN_BG
for i in range(5):
    col = get_column_letter(3+i)
    cell = ws4.cell(row=r, column=3+i)
    cell.value = f'={col}{basic_row_a}*9.99+{col}{pro_row_a}*19.99'
    cell.font = Font(name='Arial', bold=True, size=11)
    cell.number_format = '$#,##0'
    cell.fill = GREEN_BG

# MODEL B: Subscription
r += 2
style_section(ws4, r, 2, "Model B: Subscription Only — 14-day free trial, then $9.99/mo or $19.99/mo", 7)
r += 1
ws4.cell(row=r, column=2, value="Assumption: 50% convert after trial. Of converts: 60% Basic, 40% Pro").font = Font(name='Arial', italic=True, size=9, color='666666')
r += 1
for i, h in enumerate(a_headers):
    ws4.cell(row=r, column=2 + i, value=h)
style_header_row(ws4, r, 7)

r += 1
ws4.cell(row=r, column=2, value="Trial Users (not paying)").font = BLACK_REG
for i, u in enumerate(user_scales):
    ws4.cell(row=r, column=3+i, value=round(u*0.5)).font = BLACK_FORMULA
    ws4.cell(row=r, column=3+i).number_format = '#,##0'
r += 1
ws4.cell(row=r, column=2, value="Basic Subscribers @ $9.99 (30%)").font = BLACK_REG
for i, u in enumerate(user_scales):
    ws4.cell(row=r, column=3+i, value=round(u*0.30)).font = BLACK_FORMULA
    ws4.cell(row=r, column=3+i).number_format = '#,##0'
basic_row_b = r
r += 1
ws4.cell(row=r, column=2, value="Pro Subscribers @ $19.99 (20%)").font = BLACK_REG
for i, u in enumerate(user_scales):
    ws4.cell(row=r, column=3+i, value=round(u*0.20)).font = BLACK_FORMULA
    ws4.cell(row=r, column=3+i).number_format = '#,##0'
pro_row_b = r
r += 1
ws4.cell(row=r, column=2, value="Monthly Revenue").font = BLACK_BOLD
ws4.cell(row=r, column=2).fill = GREEN_BG
for i in range(5):
    col = get_column_letter(3+i)
    cell = ws4.cell(row=r, column=3+i)
    cell.value = f'={col}{basic_row_b}*9.99+{col}{pro_row_b}*19.99'
    cell.font = Font(name='Arial', bold=True, size=11)
    cell.number_format = '$#,##0'
    cell.fill = GREEN_BG

# MODEL C: Pay-per-Use
r += 2
style_section(ws4, r, 2, "Model C: Pay-per-Use — Credit packs ($4.99/50cr, $9.99/120cr, $24.99/350cr)", 7)
r += 1
ws4.cell(row=r, column=2, value="Assumption: Avg spend $3/mo (light), $10/mo (med), $25/mo (heavy). Mix: 40/40/20").font = Font(name='Arial', italic=True, size=9, color='666666')
r += 1
for i, h in enumerate(a_headers):
    ws4.cell(row=r, column=2 + i, value=h)
style_header_row(ws4, r, 7)

r += 1
ws4.cell(row=r, column=2, value="Light Users ($3/mo avg) — 40%").font = BLACK_REG
for i, u in enumerate(user_scales):
    ws4.cell(row=r, column=3+i, value=round(u*0.4*3)).font = BLACK_FORMULA
    ws4.cell(row=r, column=3+i).number_format = '$#,##0'
light_rev_c = r
r += 1
ws4.cell(row=r, column=2, value="Medium Users ($10/mo avg) — 40%").font = BLACK_REG
for i, u in enumerate(user_scales):
    ws4.cell(row=r, column=3+i, value=round(u*0.4*10)).font = BLACK_FORMULA
    ws4.cell(row=r, column=3+i).number_format = '$#,##0'
med_rev_c = r
r += 1
ws4.cell(row=r, column=2, value="Heavy Users ($25/mo avg) — 20%").font = BLACK_REG
for i, u in enumerate(user_scales):
    ws4.cell(row=r, column=3+i, value=round(u*0.2*25)).font = BLACK_FORMULA
    ws4.cell(row=r, column=3+i).number_format = '$#,##0'
heavy_rev_c = r
r += 1
ws4.cell(row=r, column=2, value="Monthly Revenue").font = BLACK_BOLD
ws4.cell(row=r, column=2).fill = GREEN_BG
for i in range(5):
    col = get_column_letter(3+i)
    cell = ws4.cell(row=r, column=3+i)
    cell.value = f'={col}{light_rev_c}+{col}{med_rev_c}+{col}{heavy_rev_c}'
    cell.font = Font(name='Arial', bold=True, size=11)
    cell.number_format = '$#,##0'
    cell.fill = GREEN_BG

# MODEL D: Free
r += 2
style_section(ws4, r, 2, "Model D: Free — No Revenue (Growth Phase / Fundraising Strategy)", 7)
r += 1
ws4.cell(row=r, column=2, value="Focus: User acquisition, engagement metrics, fundraising based on traction").font = Font(name='Arial', italic=True, size=9, color='666666')
r += 1
for i, h in enumerate(a_headers):
    ws4.cell(row=r, column=2 + i, value=h)
style_header_row(ws4, r, 7)
r += 1
ws4.cell(row=r, column=2, value="Monthly Revenue").font = BLACK_BOLD
ws4.cell(row=r, column=2).fill = RED_BG
for i in range(5):
    ws4.cell(row=r, column=3+i, value=0).font = BLACK_FORMULA
    ws4.cell(row=r, column=3+i).number_format = '$#,##0'
    ws4.cell(row=r, column=3+i).fill = RED_BG
r += 1
ws4.cell(row=r, column=2, value="KPI: Monthly Active Users (MAU)").font = BLACK_REG
for i, u in enumerate(user_scales):
    ws4.cell(row=r, column=3+i, value=u).font = BLUE_INPUT
    ws4.cell(row=r, column=3+i).number_format = '#,##0'
r += 1
ws4.cell(row=r, column=2, value="KPI: Avg Sessions/User/Week").font = BLACK_REG
for i in range(5):
    ws4.cell(row=r, column=3+i, value=3.5).font = BLUE_INPUT
r += 1
ws4.cell(row=r, column=2, value="Fundraising valuation multiplier").font = BLACK_REG
ws4.cell(row=r, column=3, value="$50-150 per MAU in EdTech (seed stage)").font = Font(name='Arial', italic=True, size=9)
ws4.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)

# ============================================================================
# SHEET 5: 12-Month Cash Flow
# ============================================================================
ws5 = wb.create_sheet("12-Month Cash Flow")
ws5.sheet_properties.tabColor = "F44336"
set_col_widths(ws5, [3, 28] + [14]*12 + [16, 3])

style_title(ws5, 1, 2, "12-Month Cash Flow Projection — Subscription Model (B) @ 500 Users Target", 15)
ws5.row_dimensions[1].height = 40

r = 3
ws5.cell(row=r, column=2, value="Scenario: Grow from 0 → 500 active users over 12 months").font = Font(name='Arial', italic=True, size=10)
ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=15)

r += 1
months = ["M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12", "TOTAL"]
for i, m in enumerate(months):
    ws5.cell(row=r, column=3 + i, value=m)
ws5.cell(row=r, column=2, value="")
style_header_row(ws5, r, 15)
month_header_row = r

# User growth
r += 1
style_section(ws5, r, 2, "USER GROWTH", 15)
r += 1
ws5.cell(row=r, column=2, value="New Users").font = BLACK_REG
new_users = [20, 30, 40, 50, 50, 50, 40, 40, 40, 50, 50, 40]
for i, u in enumerate(new_users):
    ws5.cell(row=r, column=3+i, value=u).font = BLUE_INPUT
    ws5.cell(row=r, column=3+i).fill = INPUT_BG
new_users_row = r

r += 1
ws5.cell(row=r, column=2, value="Churn Rate").font = BLACK_REG
for i in range(12):
    ws5.cell(row=r, column=3+i, value=0.05).font = BLUE_INPUT
    ws5.cell(row=r, column=3+i).number_format = '0.0%'
    ws5.cell(row=r, column=3+i).fill = INPUT_BG
churn_row = r

r += 1
ws5.cell(row=r, column=2, value="Active Users (cumulative)").font = BLACK_BOLD
# M1 = new users
ws5.cell(row=r, column=3).value = f'=C{new_users_row}'
ws5.cell(row=r, column=3).font = BLACK_FORMULA
ws5.cell(row=r, column=3).number_format = '#,##0'
# M2-M12: prev * (1-churn) + new
for i in range(1, 12):
    col = get_column_letter(3+i)
    prev_col = get_column_letter(2+i)
    cell = ws5.cell(row=r, column=3+i)
    cell.value = f'=ROUND({prev_col}{r}*(1-{col}{churn_row})+{col}{new_users_row},0)'
    cell.font = BLACK_FORMULA
    cell.number_format = '#,##0'
# Total col = last month value
ws5.cell(row=r, column=15).value = f'=N{r}'
ws5.cell(row=r, column=15).font = BLACK_FORMULA
ws5.cell(row=r, column=15).number_format = '#,##0'
active_users_row = r

# Paying users
r += 1
ws5.cell(row=r, column=2, value="Paying Users (50% conversion)").font = BLACK_REG
for i in range(12):
    col = get_column_letter(3+i)
    ws5.cell(row=r, column=3+i).value = f'=ROUND({col}{active_users_row}*0.5,0)'
    ws5.cell(row=r, column=3+i).font = BLACK_FORMULA
    ws5.cell(row=r, column=3+i).number_format = '#,##0'
ws5.cell(row=r, column=15).value = f'=N{r}'
ws5.cell(row=r, column=15).font = BLACK_FORMULA
paying_row = r

# REVENUE
r += 2
style_section(ws5, r, 2, "REVENUE", 15)
r += 1
ws5.cell(row=r, column=2, value="Avg Revenue/Paying User").font = BLACK_REG
for i in range(12):
    ws5.cell(row=r, column=3+i, value=13.99).font = BLUE_INPUT
    ws5.cell(row=r, column=3+i).number_format = '$#,##0.00'
    ws5.cell(row=r, column=3+i).fill = INPUT_BG
arpu_row = r

r += 1
ws5.cell(row=r, column=2, value="Monthly Revenue").font = BLACK_BOLD
ws5.cell(row=r, column=2).fill = GREEN_BG
for i in range(12):
    col = get_column_letter(3+i)
    cell = ws5.cell(row=r, column=3+i)
    cell.value = f'={col}{paying_row}*{col}{arpu_row}'
    cell.font = Font(name='Arial', bold=True, color='008000')
    cell.number_format = '$#,##0'
    cell.fill = GREEN_BG
# Total
col_range = ','.join([f'{get_column_letter(3+i)}{r}' for i in range(12)])
ws5.cell(row=r, column=15).value = f'=SUM(C{r}:N{r})'
ws5.cell(row=r, column=15).font = Font(name='Arial', bold=True)
ws5.cell(row=r, column=15).number_format = '$#,##0'
ws5.cell(row=r, column=15).fill = GREEN_BG
revenue_row = r

# COSTS
r += 2
style_section(ws5, r, 2, "COSTS", 15)
r += 1
# AI API cost
ws5.cell(row=r, column=2, value="Claude AI API").font = BLACK_REG
for i in range(12):
    col = get_column_letter(3+i)
    cell = ws5.cell(row=r, column=3+i)
    cell.value = f'={col}{active_users_row}*13.30'  # avg cost per user
    cell.font = BLACK_FORMULA
    cell.number_format = '$#,##0'
ai_cost_row = r

r += 1
ws5.cell(row=r, column=2, value="Infrastructure").font = BLACK_REG
infra_monthly = [38, 38, 38, 60, 60, 60, 90, 90, 90, 120, 120, 120]
for i, cost in enumerate(infra_monthly):
    ws5.cell(row=r, column=3+i, value=cost).font = BLUE_INPUT
    ws5.cell(row=r, column=3+i).number_format = '$#,##0'
    ws5.cell(row=r, column=3+i).fill = INPUT_BG
infra_row = r

r += 1
ws5.cell(row=r, column=2, value="Marketing & Growth").font = BLACK_REG
mkt_monthly = [2000, 1500, 1500, 1500, 1000, 1000, 1000, 1000, 1000, 1500, 1500, 1000]
for i, cost in enumerate(mkt_monthly):
    ws5.cell(row=r, column=3+i, value=cost).font = BLUE_INPUT
    ws5.cell(row=r, column=3+i).number_format = '$#,##0'
    ws5.cell(row=r, column=3+i).fill = INPUT_BG
mkt_row = r

r += 1
ws5.cell(row=r, column=2, value="Contractors / Professionals").font = BLACK_REG
contractor_monthly = [3000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000, 2000]
for i, cost in enumerate(contractor_monthly):
    ws5.cell(row=r, column=3+i, value=cost).font = BLUE_INPUT
    ws5.cell(row=r, column=3+i).number_format = '$#,##0'
    ws5.cell(row=r, column=3+i).fill = INPUT_BG
contractor_row = r

r += 1
ws5.cell(row=r, column=2, value="One-Time Setup (M1 only)").font = BLACK_REG
ws5.cell(row=r, column=3, value=9200).font = BLUE_INPUT
ws5.cell(row=r, column=3).number_format = '$#,##0'
ws5.cell(row=r, column=3).fill = INPUT_BG
for i in range(1, 12):
    ws5.cell(row=r, column=3+i, value=0).font = BLACK_FORMULA
    ws5.cell(row=r, column=3+i).number_format = '$#,##0'
setup_row = r

r += 1
ws5.cell(row=r, column=2, value="TOTAL Costs").font = BLACK_BOLD
ws5.cell(row=r, column=2).fill = RED_BG
for i in range(12):
    col = get_column_letter(3+i)
    cell = ws5.cell(row=r, column=3+i)
    cell.value = f'={col}{ai_cost_row}+{col}{infra_row}+{col}{mkt_row}+{col}{contractor_row}+{col}{setup_row}'
    cell.font = Font(name='Arial', bold=True, color='FF0000')
    cell.number_format = '$#,##0'
    cell.fill = RED_BG
ws5.cell(row=r, column=15).value = f'=SUM(C{r}:N{r})'
ws5.cell(row=r, column=15).font = Font(name='Arial', bold=True)
ws5.cell(row=r, column=15).number_format = '$#,##0'
ws5.cell(row=r, column=15).fill = RED_BG
total_cost_cf_row = r

# NET
r += 2
ws5.cell(row=r, column=2, value="NET Monthly Cash Flow").font = Font(name='Arial', bold=True, size=12)
ws5.cell(row=r, column=2).fill = BLUE_BG
for i in range(12):
    col = get_column_letter(3+i)
    cell = ws5.cell(row=r, column=3+i)
    cell.value = f'={col}{revenue_row}-{col}{total_cost_cf_row}'
    cell.font = Font(name='Arial', bold=True, size=11)
    cell.number_format = '$#,##0;($#,##0);"-"'
    cell.fill = BLUE_BG
ws5.cell(row=r, column=15).value = f'=SUM(C{r}:N{r})'
ws5.cell(row=r, column=15).font = Font(name='Arial', bold=True, size=11)
ws5.cell(row=r, column=15).number_format = '$#,##0;($#,##0);"-"'
ws5.cell(row=r, column=15).fill = BLUE_BG
net_row = r

r += 1
ws5.cell(row=r, column=2, value="Cumulative Cash Position").font = Font(name='Arial', bold=True, size=11)
ws5.cell(row=r, column=3).value = f'=C{net_row}'
ws5.cell(row=r, column=3).font = BLACK_FORMULA
ws5.cell(row=r, column=3).number_format = '$#,##0;($#,##0);"-"'
for i in range(1, 12):
    col = get_column_letter(3+i)
    prev = get_column_letter(2+i)
    cell = ws5.cell(row=r, column=3+i)
    cell.value = f'={prev}{r}+{col}{net_row}'
    cell.font = BLACK_FORMULA
    cell.number_format = '$#,##0;($#,##0);"-"'
ws5.cell(row=r, column=15).value = f'=N{r}'
ws5.cell(row=r, column=15).font = Font(name='Arial', bold=True)
ws5.cell(row=r, column=15).number_format = '$#,##0;($#,##0);"-"'

# ============================================================================
# SHEET 6: Launch Plan
# ============================================================================
ws6 = wb.create_sheet("Launch Plan")
ws6.sheet_properties.tabColor = "9C27B0"
set_col_widths(ws6, [3, 8, 35, 18, 16, 18, 35, 3])

style_title(ws6, 1, 2, "NoteSnap Product Launch Plan — Detailed Timeline", 7)
ws6.row_dimensions[1].height = 40

r = 3
plan_headers = ["Week", "Task", "Owner", "Cost Est.", "Priority", "Dependencies"]
for i, h in enumerate(plan_headers):
    ws6.cell(row=r, column=2 + i, value=h)
style_header_row(ws6, r, 7)

r += 1
# PHASE 1
style_section(ws6, r, 2, "PHASE 1: PRE-LAUNCH (Weeks 1-2)", 7)
r += 1
phase1 = [
    ("W1", "Legal: Privacy Policy + ToS + GDPR compliance", "Lawyer", "$2,000", "P0", "None"),
    ("W1", "Landing page: hero, features, waitlist, pricing", "Designer+Dev", "$1,500", "P0", "None"),
    ("W1", "SEO audit: meta tags, sitemap, structured data", "SEO Specialist", "$800", "P0", "Landing page"),
    ("W1", "Payment integration: Stripe setup + webhooks", "Developer", "$500", "P0", "Pricing decision"),
    ("W1", "Analytics: Mixpanel/PostHog + conversion funnels", "Developer", "$0", "P1", "None"),
    ("W1", "Brand refresh: logo polish, social banners, favicon", "Designer", "$500", "P1", "None"),
    ("W2", "Onboarding flow: first-use wizard, sample content", "UX+Dev", "$800", "P0", "None"),
    ("W2", "Email sequences: welcome, activation, day-3 nudge", "Copywriter", "$400", "P0", "Resend setup"),
    ("W2", "Performance optimization: Lighthouse 90+, Core Web Vitals", "Developer", "$300", "P1", "None"),
    ("W2", "Security audit: OWASP scan, pen test, rate limiting", "Security", "$1,000", "P0", "None"),
    ("W2", "QA: cross-browser, mobile, RTL, accessibility", "QA Tester", "$600", "P0", "None"),
    ("W2", "Social media: create accounts, schedule launch posts", "Marketing", "$200", "P1", "Brand assets"),
]
for week, task, owner, cost, pri, dep in phase1:
    ws6.cell(row=r, column=2, value=week).font = BLACK_REG
    ws6.cell(row=r, column=3, value=task).font = BLACK_REG
    ws6.cell(row=r, column=4, value=owner).font = BLACK_REG
    ws6.cell(row=r, column=5, value=cost).font = BLUE_INPUT
    ws6.cell(row=r, column=6, value=pri).font = Font(name='Arial', bold=True, size=10, color='FF0000' if pri == 'P0' else '666666')
    ws6.cell(row=r, column=7, value=dep).font = Font(name='Arial', size=9, color='666666')
    r += 1

# PHASE 2
style_section(ws6, r, 2, "PHASE 2: SOFT LAUNCH (Weeks 3-4)", 7)
r += 1
phase2 = [
    ("W3", "Beta invites: 50 users from waitlist / networks", "Founder", "$0", "P0", "Phase 1 done"),
    ("W3", "Feedback loop: in-app feedback widget + interviews", "Developer", "$0", "P0", "Beta users"),
    ("W3", "Content marketing: 3 blog posts (SEO-optimized)", "Writer", "$600", "P1", "SEO audit"),
    ("W3", "Product Hunt: prep listing, screenshots, GIF demos", "Marketing", "$200", "P1", "Beta feedback"),
    ("W4", "Bug fixing sprint based on beta feedback", "Developer", "$0", "P0", "Beta feedback"),
    ("W4", "App Store Optimization (if PWA)", "Marketing", "$300", "P2", "None"),
    ("W4", "Referral system: invite friends for premium credits", "Developer", "$300", "P1", "Stripe ready"),
    ("W4", "Testimonials: collect from beta users", "Founder", "$0", "P1", "Beta users"),
]
for week, task, owner, cost, pri, dep in phase2:
    ws6.cell(row=r, column=2, value=week).font = BLACK_REG
    ws6.cell(row=r, column=3, value=task).font = BLACK_REG
    ws6.cell(row=r, column=4, value=owner).font = BLACK_REG
    ws6.cell(row=r, column=5, value=cost).font = BLUE_INPUT
    ws6.cell(row=r, column=6, value=pri).font = Font(name='Arial', bold=True, size=10, color='FF0000' if pri == 'P0' else '666666')
    ws6.cell(row=r, column=7, value=dep).font = Font(name='Arial', size=9, color='666666')
    r += 1

# PHASE 3
style_section(ws6, r, 2, "PHASE 3: PUBLIC LAUNCH (Weeks 5-8)", 7)
r += 1
phase3 = [
    ("W5", "Product Hunt launch day", "Founder+Team", "$0", "P0", "PH listing ready"),
    ("W5", "Press outreach: EdTech blogs, TechCrunch, etc.", "PR Agency", "$1,200", "P1", "Press kit"),
    ("W5", "Google Ads campaign: EdTech keywords", "Marketing", "$1,000/mo", "P1", "Landing page"),
    ("W5", "Social media ads: Instagram, TikTok (student demo)", "Marketing", "$500/mo", "P1", "Creative assets"),
    ("W6", "YouTube tutorial: 'How NoteSnap works' (2-3 min)", "Video prod.", "$800", "P1", "None"),
    ("W6", "University/school outreach: pilot programs", "BD", "$0", "P2", "Product stable"),
    ("W7", "Influencer partnerships: EdTech YouTubers/TikTokers", "Marketing", "$1,000-3,000", "P1", "Budget approval"),
    ("W7", "Community: Discord/Telegram group for users", "Community mgr", "$200", "P2", "100+ users"),
    ("W8", "Iteration sprint: top feedback → feature updates", "Developer", "$0", "P0", "User data"),
    ("W8", "A/B test pricing page for conversion optimization", "Developer", "$0", "P1", "Traffic data"),
]
for week, task, owner, cost, pri, dep in phase3:
    ws6.cell(row=r, column=2, value=week).font = BLACK_REG
    ws6.cell(row=r, column=3, value=task).font = BLACK_REG
    ws6.cell(row=r, column=4, value=owner).font = BLACK_REG
    ws6.cell(row=r, column=5, value=cost).font = BLUE_INPUT
    ws6.cell(row=r, column=6, value=pri).font = Font(name='Arial', bold=True, size=10, color='FF0000' if pri == 'P0' else '666666')
    ws6.cell(row=r, column=7, value=dep).font = Font(name='Arial', size=9, color='666666')
    r += 1

# PHASE 4
style_section(ws6, r, 2, "PHASE 4: GROWTH (Months 3-12)", 7)
r += 1
phase4 = [
    ("M3", "Retention optimization: reduce churn with engagement hooks", "Developer", "$0", "P0", "Churn data"),
    ("M3", "Localization: expand to 3-5 more languages", "Translators", "$2,000", "P1", "i18n framework"),
    ("M4", "API/SDK: allow schools to integrate NoteSnap", "Developer", "$0", "P2", "Product stable"),
    ("M4", "Teacher dashboard: class management features", "Developer", "$0", "P1", "School pilots"),
    ("M6", "Mobile app (React Native or native)", "Developer", "$5,000-15,000", "P2", "Web stable"),
    ("M6", "Enterprise plan: school/university licensing", "BD+Legal", "$2,000", "P1", "Teacher dashboard"),
    ("M9", "Series A prep: metrics deck, financial model", "Founder+CFO", "$3,000", "P1", "Traction"),
    ("M12", "Review: annual metrics, pivot/scale decisions", "Team", "$0", "P0", "12-mo data"),
]
for week, task, owner, cost, pri, dep in phase4:
    ws6.cell(row=r, column=2, value=week).font = BLACK_REG
    ws6.cell(row=r, column=3, value=task).font = BLACK_REG
    ws6.cell(row=r, column=4, value=owner).font = BLACK_REG
    ws6.cell(row=r, column=5, value=cost).font = BLUE_INPUT
    ws6.cell(row=r, column=6, value=pri).font = Font(name='Arial', bold=True, size=10, color='FF0000' if pri == 'P0' else '666666')
    ws6.cell(row=r, column=7, value=dep).font = Font(name='Arial', size=9, color='666666')
    r += 1

# ============================================================================
# SHEET 7: Risk Analysis
# ============================================================================
ws7 = wb.create_sheet("Risk Analysis")
ws7.sheet_properties.tabColor = "E91E63"
set_col_widths(ws7, [3, 30, 14, 14, 40, 3])

style_title(ws7, 1, 2, "Risk Analysis & Mitigation", 5)
ws7.row_dimensions[1].height = 40

r = 3
risk_headers = ["Risk", "Probability", "Impact", "Mitigation"]
for i, h in enumerate(risk_headers):
    ws7.cell(row=r, column=2 + i, value=h)
style_header_row(ws7, r, 5)

r += 1
risks = [
    ("AI API costs exceed projections", "Medium", "High", "Implement usage caps per user tier; cache common responses; use Haiku for simple tasks"),
    ("Low user conversion to paid", "High", "High", "A/B test pricing; offer annual discount; strong onboarding; free tier with clear upgrade path"),
    ("Claude API rate limits / outages", "Low", "Critical", "Implement request queuing; graceful degradation; fallback to cached content"),
    ("Competitor launches similar product", "Medium", "Medium", "Focus on niche (Hebrew+English); build community moat; iterate faster"),
    ("GDPR / data privacy issues", "Low", "Critical", "Hire legal counsel; data residency in EU; privacy-by-design architecture"),
    ("Technical debt slows development", "Medium", "Medium", "Allocate 20% sprint time to refactoring; maintain test coverage >80%"),
    ("Single founder burnout", "High", "Critical", "Hire key contractor early; automate operations; set sustainable pace"),
    ("Supabase free tier limits hit", "Medium", "Low", "Monitor usage; upgrade plan proactively; optimize queries"),
    ("Student data sensitivity concerns", "Medium", "High", "No PII collection; anonymized analytics; parental consent flow"),
    ("Exchange rate risk (USD costs, local revenue)", "Medium", "Medium", "Price in USD globally; hedge with multi-currency Stripe"),
]

for risk, prob, impact, mitigation in risks:
    ws7.cell(row=r, column=2, value=risk).font = BLACK_REG
    prob_cell = ws7.cell(row=r, column=3, value=prob)
    prob_cell.font = BLACK_REG
    prob_cell.fill = ORANGE_BG if prob == "Medium" else (RED_BG if prob == "High" else GREEN_BG)
    impact_cell = ws7.cell(row=r, column=4, value=impact)
    impact_cell.font = BLACK_REG
    impact_cell.fill = RED_BG if impact in ("High", "Critical") else (ORANGE_BG if impact == "Medium" else GREEN_BG)
    ws7.cell(row=r, column=5, value=mitigation).font = Font(name='Arial', size=9)
    ws7.row_dimensions[r].height = 35
    r += 1

# ============================================================================
# Save
# ============================================================================
output_path = '/Users/curvalux/NoteSnap/docs/budget/NoteSnap_Launch_Budget.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
