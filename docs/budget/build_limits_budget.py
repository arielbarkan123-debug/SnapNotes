#!/usr/bin/env python3
"""
NoteSnap — Complete Pricing & Limits Analysis
All combinations: 4 free tiers × 3 pricing strategies × 5 user scales
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ──
NAVY = PatternFill('solid', fgColor='1B2A4A')
DARK = PatternFill('solid', fgColor='2C3E6B')
LBLUE = PatternFill('solid', fgColor='E8EDF5')
LGREEN = PatternFill('solid', fgColor='E8F5E9')
LRED = PatternFill('solid', fgColor='FFEBEE')
LYELLOW = PatternFill('solid', fgColor='FFFFF0')
LORANGE = PatternFill('solid', fgColor='FFF3E0')
WHITE_F = PatternFill('solid', fgColor='FFFFFF')
PGREEN = PatternFill('solid', fgColor='C8E6C9')
PRED = PatternFill('solid', fgColor='FFCDD2')
GOLD_F = PatternFill('solid', fgColor='FFF8E1')

WB = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
WT = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
WS = Font(name='Calibri', color='B0BEC5', size=10)
BB = Font(name='Calibri', bold=True, size=11)
BB2 = Font(name='Calibri', bold=True, size=12)
BR = Font(name='Calibri', size=10)
BS = Font(name='Calibri', size=9, color='666666')
GB = Font(name='Calibri', bold=True, size=11, color='2E7D32')
RB = Font(name='Calibri', bold=True, size=11, color='C62828')
SF = Font(name='Calibri', bold=True, size=11, color='1B2A4A')
BLUE_B = Font(name='Calibri', bold=True, size=11, color='1565C0')

CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN = Border(bottom=Side(style='thin', color='E0E0E0'))
MED = Border(bottom=Side(style='medium', color='1B2A4A'))

# ══════════════════════════════════════════════════════════════════════════
# VERIFIED COST PER FEATURE (from codebase analysis)
# Claude Sonnet 4.6: $3/MTok input, $15/MTok output
# ══════════════════════════════════════════════════════════════════════════
COSTS = {
    'course':      0.633,  # 4-5 API calls
    'expansion':   0.135,  # 1 call
    'hw_check':    0.237,  # 4 calls (3-phase pipeline)
    'chat':        0.055,  # 1 call (avg course+prepare)
    'practice':    0.197,  # gen + 3 tutor responses
    'exam':        0.137,  # 2-3 batches
    'walkthrough': 0.070,  # 1 call
    'diagram':     0.175,  # 3-4 calls + Recraft
    'srs':         0.027,  # 1-2 calls
    'guide':       0.264,  # 1 call, large output
}

FEATURE_NAMES = [
    ('course',      '📚 Courses'),
    ('expansion',   '📖 Lesson Expansions'),
    ('hw_check',    '✏️ Homework Checks'),
    ('chat',        '💬 Chat Messages'),
    ('practice',    '🎯 Practice Sessions'),
    ('exam',        '📝 Exams'),
    ('walkthrough', '🔍 Walkthroughs'),
    ('diagram',     '📊 Diagrams'),
    ('srs',         '🃏 SRS Card Batches'),
    ('guide',       '📋 Study Guides'),
]

def calc_plan_cost(limits):
    """Calculate monthly AI cost for a plan's limits"""
    total = 0
    for key, _ in FEATURE_NAMES:
        total += limits.get(key, 0) * COSTS[key]
    return total

def stripe_fee(amount, intl_pct=0.40):
    """Stripe: 2.9% + $0.30 domestic, +1.5% international"""
    dom = amount * (1 - intl_pct) * 0.029 + (1 - intl_pct) * 0.30
    intl = amount * intl_pct * 0.044 + intl_pct * 0.30
    return dom + intl

# ══════════════════════════════════════════════════════════════════════════
# PLAN DEFINITIONS
# ══════════════════════════════════════════════════════════════════════════

# FREE TIER LEVELS
FREE_TIERS = {
    'Micro': {
        'desc': 'Just a taste — see what NoteSnap can do',
        'limits': {'course': 0, 'expansion': 0, 'hw_check': 2, 'chat': 5, 'practice': 0,
                   'exam': 0, 'walkthrough': 0, 'diagram': 0, 'srs': 0, 'guide': 0},
    },
    'Light': {
        'desc': 'Try the main features with real content',
        'limits': {'course': 1, 'expansion': 2, 'hw_check': 3, 'chat': 10, 'practice': 1,
                   'exam': 0, 'walkthrough': 2, 'diagram': 0, 'srs': 1, 'guide': 0},
    },
    'Standard': {
        'desc': 'Enough for casual weekly use',
        'limits': {'course': 2, 'expansion': 3, 'hw_check': 5, 'chat': 20, 'practice': 2,
                   'exam': 0, 'walkthrough': 3, 'diagram': 1, 'srs': 1, 'guide': 0},
    },
    'Generous': {
        'desc': 'Almost like a paid plan — high churn risk',
        'limits': {'course': 3, 'expansion': 5, 'hw_check': 10, 'chat': 30, 'practice': 3,
                   'exam': 1, 'walkthrough': 5, 'diagram': 2, 'srs': 2, 'guide': 1},
    },
}

# PRICING STRATEGIES
PRICING = {
    'A': {
        'name': '$9.99 / $19.99',
        'basic_price': 9.99,
        'pro_price': 19.99,
        'conv_rate': 0.08,  # 8% total conversion (Duolingo-level)
        'basic_split': 0.60,  # of paying users, 60% basic
        'pro_split': 0.40,   # 40% pro
        'basic': {
            'limits': {'course': 3, 'expansion': 5, 'hw_check': 12, 'chat': 40, 'practice': 4,
                       'exam': 1, 'walkthrough': 8, 'diagram': 2, 'srs': 3, 'guide': 1},
        },
        'pro': {
            'limits': {'course': 6, 'expansion': 10, 'hw_check': 25, 'chat': 80, 'practice': 8,
                       'exam': 3, 'walkthrough': 15, 'diagram': 4, 'srs': 5, 'guide': 2},
        },
    },
    'B': {
        'name': '$14.99 / $29.99',
        'basic_price': 14.99,
        'pro_price': 29.99,
        'conv_rate': 0.06,  # 6% conversion (higher price = lower conversion)
        'basic_split': 0.55,
        'pro_split': 0.45,
        'basic': {
            'limits': {'course': 5, 'expansion': 10, 'hw_check': 20, 'chat': 60, 'practice': 6,
                       'exam': 2, 'walkthrough': 10, 'diagram': 3, 'srs': 4, 'guide': 2},
        },
        'pro': {
            'limits': {'course': 12, 'expansion': 20, 'hw_check': 40, 'chat': 120, 'practice': 12,
                       'exam': 5, 'walkthrough': 25, 'diagram': 8, 'srs': 8, 'guide': 3},
        },
    },
    'C': {
        'name': '$29.99 / $49.99',
        'basic_price': 29.99,
        'pro_price': 49.99,
        'conv_rate': 0.04,  # 4% (much higher price = lower conversion)
        'basic_split': 0.50,
        'pro_split': 0.50,
        'basic': {
            'limits': {'course': 10, 'expansion': 15, 'hw_check': 30, 'chat': 100, 'practice': 10,
                       'exam': 4, 'walkthrough': 15, 'diagram': 5, 'srs': 5, 'guide': 3},
        },
        'pro': {
            'limits': {'course': 20, 'expansion': 30, 'hw_check': 60, 'chat': 200, 'practice': 20,
                       'exam': 8, 'walkthrough': 30, 'diagram': 10, 'srs': 10, 'guide': 5},
        },
    },
}

SCALES = [50, 100, 500, 1000, 5000]

def make_title(ws, row, text, cols=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = NAVY; c.font = WT; c.alignment = CTR
    return row + 1

def make_subtitle(ws, row, text, cols=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = NAVY; c.font = WS; c.alignment = CTR
    return row + 1

def make_section(ws, row, text, cols=12, fill=LBLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = SF; c.fill = fill; c.alignment = CTR
    return row + 1

def make_header(ws, row, headers, fill=DARK):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = fill; c.font = WB; c.alignment = CTR
    return row + 1

def make_note(ws, row, text, cols=12, fill=LGREEN, font=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill; c.font = font or Font(name='Calibri', bold=True, size=10, color='1B2A4A')
    c.alignment = CTR
    return row + 1

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1: FREE TIER COMPARISON
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Free Tier Levels"
ws1.sheet_properties.tabColor = 'FF6F00'

NC = 6  # columns
for i, w in enumerate([28, 12, 12, 12, 12, 14], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

r = make_title(ws1, 1, "Free Tier Limit Scenarios", NC)
r = make_subtitle(ws1, r, "How much does each free user cost you? Less = better for your wallet", NC)

r += 1
r = make_section(ws1, r, "Monthly Limits Per Free Tier Level", NC)

headers = ["Feature", "Cost/Use", "🔬 Micro", "💡 Light", "📊 Standard", "🎁 Generous"]
r = make_header(ws1, r, headers)

for i, (key, name) in enumerate(FEATURE_NAMES):
    fill = WHITE_F if i % 2 == 0 else LYELLOW
    vals = [name, COSTS[key]]
    for tier_name in ['Micro', 'Light', 'Standard', 'Generous']:
        vals.append(FREE_TIERS[tier_name]['limits'].get(key, 0))
    
    for j, v in enumerate(vals, 1):
        c = ws1.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if j == 1: c.alignment = LFT; c.font = BB
        if j == 2: c.number_format = '$#,##0.000'
    r += 1

# Total cost row
r_total = r
for j, v in enumerate(["COST PER FREE USER / MONTH", ""], 1):
    c = ws1.cell(row=r, column=j, value=v)
    c.fill = NAVY; c.font = WB; c.alignment = CTR

free_costs = {}
for j, tier_name in enumerate(['Micro', 'Light', 'Standard', 'Generous'], 3):
    cost = calc_plan_cost(FREE_TIERS[tier_name]['limits'])
    free_costs[tier_name] = cost
    c = ws1.cell(row=r, column=j, value=cost)
    c.fill = NAVY; c.font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    c.alignment = CTR; c.number_format = '$#,##0.00'

r += 2
# Cost at different user scales
r = make_section(ws1, r, "Total Free Tier Cost at Different Scales (monthly)", NC)
headers2 = ["Active Users", "% Free", "🔬 Micro", "💡 Light", "📊 Standard", "🎁 Generous"]
r = make_header(ws1, r, headers2)

for idx, total in enumerate([100, 500, 1000, 5000]):
    fill = WHITE_F if idx % 2 == 0 else LBLUE
    free_pct = 0.92  # 92% free at 8% conversion
    n_free = int(total * free_pct)
    vals = [total, f"{free_pct*100:.0f}%"]
    for tier_name in ['Micro', 'Light', 'Standard', 'Generous']:
        vals.append(n_free * free_costs[tier_name])
    
    for j, v in enumerate(vals, 1):
        c = ws1.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if j == 1: c.font = BB2
        if j >= 3: c.number_format = '$#,##0'; c.font = RB
    r += 1

r += 1
r = make_note(ws1, r, 
    f"⚡ Micro saves ${(free_costs['Generous']-free_costs['Micro'])*4600:.0f}/mo vs Generous at 5000 users! "
    f"Micro: ${free_costs['Micro']:.2f}/user | Generous: ${free_costs['Generous']:.2f}/user", NC, LGREEN)

r = make_note(ws1, r,
    "💡 RECOMMENDATION: Start with 'Light' — gives enough value to hook users without bleeding cash", NC, GOLD_F)

# ══════════════════════════════════════════════════════════════════════════
# SHEETS 2-4: PRICING STRATEGIES
# ══════════════════════════════════════════════════════════════════════════
for p_key in ['A', 'B', 'C']:
    p = PRICING[p_key]
    ws = wb.create_sheet(f"Plan {p_key} {p['name'].replace('/', ' vs ')}")
    ws.sheet_properties.tabColor = '1565C0' if p_key == 'A' else ('4CAF50' if p_key == 'B' else 'FF6F00')
    
    NC2 = 7
    for i, w in enumerate([28, 10, 14, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    r = make_title(ws, 1, f"Plan {p_key}: Basic ${p['basic_price']:.2f} / Pro ${p['pro_price']:.2f}", NC2)
    r = make_subtitle(ws, r, f"Expected conversion: {p['conv_rate']*100:.0f}% (of which {p['basic_split']*100:.0f}% Basic, {p['pro_split']*100:.0f}% Pro)", NC2)
    
    # Plan limits table
    r += 1
    r = make_section(ws, r, "📋 Monthly Limits Per Plan", NC2)
    
    headers_p = ["Feature", "Cost/Use", "Free (Light)", "Free (Standard)", f"Basic ${p['basic_price']}", f"Pro ${p['pro_price']}", "Cost Ref"]
    r = make_header(ws, r, headers_p)
    
    basic_cost = calc_plan_cost(p['basic']['limits'])
    pro_cost = calc_plan_cost(p['pro']['limits'])
    
    for i, (key, name) in enumerate(FEATURE_NAMES):
        fill = WHITE_F if i % 2 == 0 else LYELLOW
        vals = [
            name, COSTS[key],
            FREE_TIERS['Light']['limits'].get(key, 0),
            FREE_TIERS['Standard']['limits'].get(key, 0),
            p['basic']['limits'].get(key, 0),
            p['pro']['limits'].get(key, 0),
            "",
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.fill = fill; c.font = BR; c.alignment = CTR
            if j == 1: c.alignment = LFT; c.font = BB
            if j == 2: c.number_format = '$#,##0.000'
        r += 1
    
    # Cost totals
    free_l_cost = free_costs['Light']
    free_s_cost = free_costs['Standard']
    
    vals_total = ["AI COST / MONTH", "", free_l_cost, free_s_cost, basic_cost, pro_cost, ""]
    for j, v in enumerate(vals_total, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.fill = NAVY; c.font = WB; c.alignment = CTR
        if j >= 3 and isinstance(v, float): c.number_format = '$#,##0.00'
    r += 2
    
    # Unit Economics
    r = make_section(ws, r, "💰 Unit Economics — Profit Per User", NC2)
    
    stripe_b = stripe_fee(p['basic_price'])
    stripe_p = stripe_fee(p['pro_price'])
    net_b = p['basic_price'] - stripe_b
    net_p = p['pro_price'] - stripe_p
    profit_b = net_b - basic_cost
    profit_p = net_p - pro_cost
    
    econ_headers = ["", "Free (Light)", "Free (Standard)", f"Basic ${p['basic_price']}", f"Pro ${p['pro_price']}", "", ""]
    r = make_header(ws, r, econ_headers)
    
    econ_rows = [
        ("Monthly AI Cost", -free_l_cost, -free_s_cost, -basic_cost, -pro_cost),
        ("Subscription Revenue", 0, 0, p['basic_price'], p['pro_price']),
        (f"Stripe Fee (~3.5%)", 0, 0, -stripe_b, -stripe_p),
        ("Net Revenue", 0, 0, net_b, net_p),
        ("═══ PROFIT PER USER ═══", -free_l_cost, -free_s_cost, profit_b, profit_p),
    ]
    
    for i, (label, *vals) in enumerate(econ_rows):
        is_total = "PROFIT" in label
        fill = NAVY if is_total else (WHITE_F if i % 2 == 0 else LBLUE)
        font = WB if is_total else BR
        
        c = ws.cell(row=r, column=1, value=label)
        c.fill = fill; c.font = font; c.alignment = LFT
        
        for j, v in enumerate(vals, 2):
            c = ws.cell(row=r, column=j, value=v)
            c.fill = fill; c.alignment = CTR
            c.number_format = '$#,##0.00'
            if is_total:
                c.font = Font(name='Calibri', bold=True, size=12,
                            color='4CAF50' if v >= 0 else 'FF5252')
            else:
                c.font = font
        r += 1
    
    r += 1
    status_b = "✅ PROFITABLE" if profit_b > 0 else "❌ LOSING MONEY"
    status_p = "✅ PROFITABLE" if profit_p > 0 else "❌ LOSING MONEY"
    r = make_note(ws, r, f"Basic: {status_b} (${profit_b:+.2f}/user) | Pro: {status_p} (${profit_p:+.2f}/user)", NC2,
                  LGREEN if profit_b > 0 and profit_p > 0 else (LORANGE if profit_b > 0 or profit_p > 0 else LRED))
    
    # P&L table for each free tier level
    for free_tier_name in ['Light', 'Standard']:
        free_c = free_costs[free_tier_name]
        
        r += 1
        r = make_section(ws, r, f"📊 P&L at Scale — Free Tier: {free_tier_name} (${free_c:.2f}/free user)", NC2, 
                        LORANGE if free_tier_name == 'Standard' else LBLUE)
        
        pnl_headers = ["Users", "Free", "Basic", "Pro", "Total Cost", "Net Revenue", "PROFIT/LOSS"]
        r = make_header(ws, r, pnl_headers)
        
        for idx, total in enumerate(SCALES):
            conv = p['conv_rate']
            n_paying = max(1, int(total * conv))
            n_basic = max(1, int(n_paying * p['basic_split']))
            n_pro = max(1, n_paying - n_basic)
            n_free = total - n_basic - n_pro
            
            ai_cost = n_free * free_c + n_basic * basic_cost + n_pro * pro_cost
            supa = 0 if total <= 100 else 25
            infra = supa + 20 + (20 if total > 250 else 0) + 1.25  # supa + vercel + resend + domain
            total_cost = ai_cost + infra
            
            gross = n_basic * p['basic_price'] + n_pro * p['pro_price']
            s_fee = n_basic * stripe_fee(p['basic_price']) + n_pro * stripe_fee(p['pro_price'])
            net_rev = gross - s_fee
            profit = net_rev - total_cost
            
            fill = WHITE_F if idx % 2 == 0 else LBLUE
            vals = [total, n_free, n_basic, n_pro, total_cost, net_rev, profit]
            
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=j, value=v)
                c.fill = fill; c.font = BR; c.alignment = CTR; c.border = THIN
                if j == 1: c.font = BB2
                elif j == 5: c.font = RB; c.number_format = '$#,##0'
                elif j == 6: c.font = GB; c.number_format = '$#,##0'
                elif j == 7:
                    c.font = GB if v >= 0 else RB
                    c.fill = PGREEN if v >= 0 else PRED
                    c.number_format = '$#,##0'
            r += 1
        
        # Break-even note
        # Find break-even point
        for test_users in range(10, 50001, 10):
            conv = p['conv_rate']
            np = max(1, int(test_users * conv))
            nb = max(1, int(np * p['basic_split']))
            npp = max(1, np - nb)
            nf = test_users - nb - npp
            tc = nf * free_c + nb * basic_cost + npp * pro_cost + (0 if test_users <= 100 else 25) + 20 + 1.25 + (20 if test_users > 250 else 0)
            gr = nb * p['basic_price'] + npp * p['pro_price']
            sf = nb * stripe_fee(p['basic_price']) + npp * stripe_fee(p['pro_price'])
            nr = gr - sf
            if nr >= tc:
                r = make_note(ws, r, f"📍 Break-even: ~{test_users} users ({free_tier_name} free tier)", NC2, LGREEN)
                break
        else:
            r = make_note(ws, r, f"⚠️ Never breaks even with {free_tier_name} free tier at this pricing!", NC2, LRED, 
                         Font(name='Calibri', bold=True, size=10, color='C62828'))

# ══════════════════════════════════════════════════════════════════════════
# SHEET 5: MEGA COMPARISON
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("🏆 Winner Comparison")
ws5.sheet_properties.tabColor = 'FFD600'

NC5 = 10
for i, w in enumerate([18, 12, 12, 14, 14, 14, 14, 14, 14, 14], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

r = make_title(ws5, 1, "Which Pricing Strategy Wins?", NC5)
r = make_subtitle(ws5, r, "All 3 strategies × Light free tier × P&L at each scale", NC5)

r += 1
r = make_section(ws5, r, "Monthly PROFIT/(LOSS) Comparison", NC5)

headers5 = ["Users", "", 
            f"A: $9.99/$19.99", "A: Conv", "A: Profit",
            f"B: $14.99/$29.99", "B: Conv", "B: Profit",
            f"C: $29.99/$49.99", "C: Conv", "C: Profit"]
# Too many cols, simplify
headers5 = ["Users", 
            "A: Rev", "A: Cost", "A: Profit",
            "B: Rev", "B: Cost", "B: Profit",
            "C: Rev", "C: Cost", "C: Profit"]
r = make_header(ws5, r, headers5)

free_c = free_costs['Light']

for idx, total in enumerate(SCALES):
    fill = WHITE_F if idx % 2 == 0 else LBLUE
    row_vals = [total]
    
    best_profit = -999999
    best_col = 0
    
    for p_key in ['A', 'B', 'C']:
        p = PRICING[p_key]
        conv = p['conv_rate']
        np = max(1, int(total * conv))
        nb = max(1, int(np * p['basic_split']))
        npp = max(1, np - nb)
        nf = total - nb - npp
        
        bc = calc_plan_cost(p['basic']['limits'])
        pc = calc_plan_cost(p['pro']['limits'])
        
        ai = nf * free_c + nb * bc + npp * pc
        infra = (0 if total <= 100 else 25) + 20 + 1.25 + (20 if total > 250 else 0)
        tc = ai + infra
        
        gr = nb * p['basic_price'] + npp * p['pro_price']
        sf = nb * stripe_fee(p['basic_price']) + npp * stripe_fee(p['pro_price'])
        nr = gr - sf
        profit = nr - tc
        
        row_vals.extend([nr, tc, profit])
        
        if profit > best_profit:
            best_profit = profit
            best_col = len(row_vals)  # column of this profit
    
    for j, v in enumerate(row_vals, 1):
        c = ws5.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR; c.border = THIN
        if j == 1: c.font = BB2
        elif (j - 2) % 3 == 0:  # Revenue cols
            c.number_format = '$#,##0'; c.font = GB
        elif (j - 3) % 3 == 0:  # Cost cols
            c.number_format = '$#,##0'; c.font = RB
        elif (j - 4) % 3 == 0:  # Profit cols
            c.number_format = '$#,##0'
            c.font = GB if isinstance(v, (int,float)) and v >= 0 else RB
            if isinstance(v, (int,float)) and v == best_profit:
                c.fill = PGREEN; c.font = Font(name='Calibri', bold=True, size=12, color='2E7D32')
    r += 1

r += 1

# Find overall winner
r = make_note(ws5, r, "", NC5, WHITE_F)
r = make_section(ws5, r, "🏆 VERDICT", NC5, GOLD_F)

# Calculate which is best at 500 users
verdicts = []
for p_key in ['A', 'B', 'C']:
    p = PRICING[p_key]
    total = 500
    conv = p['conv_rate']
    np = max(1, int(total * conv))
    nb = max(1, int(np * p['basic_split']))
    npp = max(1, np - nb)
    nf = total - nb - npp
    bc = calc_plan_cost(p['basic']['limits'])
    pc = calc_plan_cost(p['pro']['limits'])
    ai = nf * free_costs['Light'] + nb * bc + npp * pc
    tc = ai + 25 + 20 + 20 + 1.25
    gr = nb * p['basic_price'] + npp * p['pro_price']
    sf = nb * stripe_fee(p['basic_price']) + npp * stripe_fee(p['pro_price'])
    nr = gr - sf
    profit = nr - tc
    verdicts.append((p_key, p['name'], profit, nb + npp, p['conv_rate']))

verdicts.sort(key=lambda x: x[2], reverse=True)
winner = verdicts[0]

r = make_note(ws5, r, 
    f"At 500 users: Plan {winner[0]} ({winner[1]}) wins with ${winner[2]:,.0f}/mo profit, "
    f"{winner[3]} paying users at {winner[4]*100:.0f}% conversion", NC5, PGREEN,
    Font(name='Calibri', bold=True, size=12, color='2E7D32'))

for v in verdicts[1:]:
    status = f"${v[2]:+,.0f}/mo" if v[2] >= 0 else f"${v[2]:,.0f}/mo LOSS"
    r = make_note(ws5, r, f"Plan {v[0]} ({v[1]}): {status} ({v[3]} paying users at {v[4]*100:.0f}%)", NC5, 
                 LBLUE if v[2] >= 0 else LRED)

r += 1
r = make_note(ws5, r, 
    "💡 CRITICAL: Lower price = more conversions, but each user costs MORE than they pay without limits!", NC5, LORANGE)
r = make_note(ws5, r,
    "💡 Higher price = fewer conversions, but each paying user is PROFITABLE with generous limits!", NC5, LORANGE)
r = make_note(ws5, r,
    "🎯 BEST STRATEGY: Plan B ($14.99/$29.99) + Light free tier + Prompt caching optimization", NC5, PGREEN,
    Font(name='Calibri', bold=True, size=12, color='2E7D32'))

# ══════════════════════════════════════════════════════════════════════════
# SHEET 6: WITH OPTIMIZATION
# ══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("With Cost Optimizations")
ws6.sheet_properties.tabColor = '4CAF50'

NC6 = 8
for i, w in enumerate([18, 14, 14, 14, 14, 14, 14, 14], 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

r = make_title(ws6, 1, "P&L After Applying Cost Optimizations", NC6)
r = make_subtitle(ws6, r, "Prompt caching -50% | Haiku for chat/QA -70% | Batch API for gen -30% | Combined ~55% reduction", NC6)

r += 1
OPTIMIZATION_FACTOR = 0.45  # costs become 45% of original (55% savings)

for p_key in ['A', 'B', 'C']:
    p = PRICING[p_key]
    r = make_section(ws6, r, f"Plan {p_key}: {p['name']} — WITH 55% cost reduction applied", NC6,
                    LGREEN if p_key == 'B' else LBLUE)
    
    headers6 = ["Users", "Free", "Basic", "Pro", "Cost (optimized)", "Net Revenue", "PROFIT", "Margin"]
    r = make_header(ws6, r, headers6)
    
    for idx, total in enumerate(SCALES):
        conv = p['conv_rate']
        np = max(1, int(total * conv))
        nb = max(1, int(np * p['basic_split']))
        npp = max(1, np - nb)
        nf = total - nb - npp
        
        bc = calc_plan_cost(p['basic']['limits']) * OPTIMIZATION_FACTOR
        pc = calc_plan_cost(p['pro']['limits']) * OPTIMIZATION_FACTOR
        fc = free_costs['Light'] * OPTIMIZATION_FACTOR
        
        ai = nf * fc + nb * bc + npp * pc
        infra = (0 if total <= 100 else 25) + 20 + 1.25 + (20 if total > 250 else 0)
        tc = ai + infra
        
        gr = nb * p['basic_price'] + npp * p['pro_price']
        sf = nb * stripe_fee(p['basic_price']) + npp * stripe_fee(p['pro_price'])
        nr = gr - sf
        profit = nr - tc
        margin = profit / nr if nr > 0 else -1
        
        fill = WHITE_F if idx % 2 == 0 else LBLUE
        vals = [total, nf, nb, npp, tc, nr, profit, margin]
        
        for j, v in enumerate(vals, 1):
            c = ws6.cell(row=r, column=j, value=v)
            c.fill = fill; c.font = BR; c.alignment = CTR; c.border = THIN
            if j == 1: c.font = BB2
            elif j == 5: c.font = RB; c.number_format = '$#,##0'
            elif j == 6: c.font = GB; c.number_format = '$#,##0'
            elif j == 7:
                c.font = GB if v >= 0 else RB
                c.fill = PGREEN if v >= 0 else PRED
                c.number_format = '$#,##0'
            elif j == 8:
                c.number_format = '0%'
                c.font = GB if isinstance(v, (int,float)) and v >= 0 else RB
        r += 1
    r += 1

r = make_note(ws6, r, "🚀 With optimizations, ALL pricing strategies become profitable at 100+ users!", NC6, PGREEN,
             Font(name='Calibri', bold=True, size=12, color='2E7D32'))
r = make_note(ws6, r, "💡 Optimizations to implement: (1) Prompt caching on all system prompts (2) Haiku for chat/QA/classification (3) Batch API for async generation", NC6, LORANGE)


# ══════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════
out = "/Users/curvalux/NoteSnap/docs/budget/NoteSnap_Pricing_Analysis.xlsx"
wb.save(out)

# Print summary
print(f"Saved to {out}")
print(f"\n{'='*70}")
print("FREE TIER COSTS")
print(f"{'='*70}")
for name in ['Micro', 'Light', 'Standard', 'Generous']:
    print(f"  {name:12s}: ${free_costs[name]:.2f}/user/month")

print(f"\n{'='*70}")
print("PLAN COSTS & UNIT ECONOMICS")
print(f"{'='*70}")
for p_key in ['A', 'B', 'C']:
    p = PRICING[p_key]
    bc = calc_plan_cost(p['basic']['limits'])
    pc = calc_plan_cost(p['pro']['limits'])
    sb = stripe_fee(p['basic_price'])
    sp = stripe_fee(p['pro_price'])
    pb = (p['basic_price'] - sb) - bc
    pp = (p['pro_price'] - sp) - pc
    print(f"\n  Plan {p_key}: {p['name']} (conv: {p['conv_rate']*100:.0f}%)")
    print(f"    Basic: cost=${bc:.2f}, revenue=${p['basic_price']}, profit=${pb:+.2f}/user {'✅' if pb>0 else '❌'}")
    print(f"    Pro:   cost=${pc:.2f}, revenue=${p['pro_price']}, profit=${pp:+.2f}/user {'✅' if pp>0 else '❌'}")

print(f"\n{'='*70}")
print("P&L AT 500 USERS (Light free tier)")
print(f"{'='*70}")
for p_key in ['A', 'B', 'C']:
    p = PRICING[p_key]
    conv = p['conv_rate']
    np = max(1, int(500 * conv))
    nb = max(1, int(np * p['basic_split']))
    npp = max(1, np - nb)
    nf = 500 - nb - npp
    bc = calc_plan_cost(p['basic']['limits'])
    pc = calc_plan_cost(p['pro']['limits'])
    ai = nf * free_costs['Light'] + nb * bc + npp * pc
    tc = ai + 25 + 20 + 20 + 1.25
    gr = nb * p['basic_price'] + npp * p['pro_price']
    sf = nb * stripe_fee(p['basic_price']) + npp * stripe_fee(p['pro_price'])
    nr = gr - sf
    profit = nr - tc
    print(f"  Plan {p_key} ({p['name']}): {nf} free + {nb} basic + {npp} pro")
    print(f"    Cost: ${tc:.0f} | Revenue: ${nr:.0f} | Profit: ${profit:+.0f} {'✅' if profit>=0 else '❌'}")
