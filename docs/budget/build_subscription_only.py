#!/usr/bin/env python3
"""
NoteSnap — Pure Subscription Model (No Free Tier)
"We replace private tutors at 90% less"
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ──
NAVY = PatternFill('solid', fgColor='1B2A4A')
DARK = PatternFill('solid', fgColor='2C3E6B')
GOLD_BG = PatternFill('solid', fgColor='F9F3E3')
LBLUE = PatternFill('solid', fgColor='E8EDF5')
LGREEN = PatternFill('solid', fgColor='E8F5E9')
LRED = PatternFill('solid', fgColor='FFEBEE')
LORANGE = PatternFill('solid', fgColor='FFF3E0')
WHITE_F = PatternFill('solid', fgColor='FFFFFF')
PGREEN = PatternFill('solid', fgColor='C8E6C9')
PRED = PatternFill('solid', fgColor='FFCDD2')
GOLD_H = PatternFill('solid', fgColor='FFF8E1')
PURPLE = PatternFill('solid', fgColor='F3E5F5')
DARK_GOLD = PatternFill('solid', fgColor='5D4E37')

WB = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
WT = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
WS = Font(name='Calibri', color='B0BEC5', size=10)
BB = Font(name='Calibri', bold=True, size=11)
BB2 = Font(name='Calibri', bold=True, size=12)
BB3 = Font(name='Calibri', bold=True, size=14)
BR = Font(name='Calibri', size=10)
GB = Font(name='Calibri', bold=True, size=11, color='2E7D32')
GB2 = Font(name='Calibri', bold=True, size=13, color='2E7D32')
GB3 = Font(name='Calibri', bold=True, size=16, color='2E7D32')
RB = Font(name='Calibri', bold=True, size=11, color='C62828')
SF = Font(name='Calibri', bold=True, size=11, color='1B2A4A')
GOLD_F = Font(name='Calibri', bold=True, size=11, color='B8860B')
GOLD_BIG = Font(name='Calibri', bold=True, size=14, color='B8860B')
WHITE_BIG = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN = Border(bottom=Side(style='thin', color='E0E0E0'))
MED_B = Border(bottom=Side(style='medium', color='1B2A4A'))

# ── VERIFIED COSTS (Claude Sonnet 4.6, multi-call pipelines) ──
C = {
    'course':    0.633,  'expansion': 0.135,  'hw_check': 0.237,
    'chat':      0.055,  'practice':  0.197,  'exam':     0.137,
    'walkthrough': 0.070, 'diagram':  0.175,  'srs':      0.027,
    'guide':     0.264,
}

FEATURES = [
    ('course',      'Course Generations'),
    ('expansion',   'Lesson Expansions'),
    ('hw_check',    'Homework Checks'),
    ('chat',        'AI Chat Messages'),
    ('practice',    'Practice Sessions'),
    ('exam',        'Mock Exams'),
    ('walkthrough', 'Step Walkthroughs'),
    ('diagram',     'Visual Diagrams'),
    ('srs',         'SRS Review Batches'),
    ('guide',       'Study Guides'),
]

OPT = 0.45  # 55% cost reduction with caching+haiku+batch

def calc_cost(usage, opt=1.0):
    return sum(usage.get(k,0) * v for k,v in C.items()) * opt

def stripe(amount, intl=0.40):
    dom = amount * (1-intl) * 0.029 + (1-intl) * 0.30
    intl_f = amount * intl * 0.044 + intl * 0.30
    return dom + intl_f

def infra(users):
    s = 0 if users <= 100 else (25 if users <= 5000 else 599)
    return s + 20 + (20 if users > 250 else 0) + 1.25

SCALES = [50, 100, 500, 1000, 5000]

def mt(ws, r, text, nc):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(row=r, column=1, value=text)
    c.fill = NAVY; c.font = WT; c.alignment = CTR
    return r+1

def ms(ws, r, text, nc):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(row=r, column=1, value=text)
    c.fill = NAVY; c.font = WS; c.alignment = CTR
    return r+1

def sec(ws, r, text, nc, fill=LBLUE):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(row=r, column=1, value=text)
    c.font = SF; c.fill = fill; c.alignment = CTR
    return r+1

def hdr(ws, r, headers, fill=DARK):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=j, value=h)
        c.fill = fill; c.font = WB; c.alignment = CTR; c.border = MED_B
    return r+1

def nt(ws, r, text, nc, fill=LGREEN, font=None):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(row=r, column=1, value=text)
    c.fill = fill; c.font = font or Font(name='Calibri', bold=True, size=10, color='1B2A4A')
    c.alignment = CTR
    return r+1

# ══════════════════════════════════════════════════════════════════
# PLAN DEFINITIONS — generous limits since EVERYONE PAYS
# ══════════════════════════════════════════════════════════════════

# Usage = what students actually USE (not limits — limits are higher)
PLANS = {
    'Student': {
        'usage': {
            'course': 5, 'expansion': 10, 'hw_check': 20, 'chat': 50,
            'practice': 5, 'exam': 2, 'walkthrough': 10, 'diagram': 3,
            'srs': 3, 'guide': 1,
        },
        'limits': {
            'course': 8, 'expansion': 15, 'hw_check': 30, 'chat': 80,
            'practice': 8, 'exam': 3, 'walkthrough': 15, 'diagram': 5,
            'srs': 5, 'guide': 2,
        },
    },
    'Pro': {
        'usage': {
            'course': 10, 'expansion': 20, 'hw_check': 40, 'chat': 100,
            'practice': 10, 'exam': 5, 'walkthrough': 20, 'diagram': 8,
            'srs': 8, 'guide': 3,
        },
        'limits': {
            'course': 15, 'expansion': 30, 'hw_check': 60, 'chat': 150,
            'practice': 15, 'exam': 8, 'walkthrough': 30, 'diagram': 12,
            'srs': 12, 'guide': 5,
        },
    },
    'Family': {
        'usage': {
            'course': 15, 'expansion': 30, 'hw_check': 60, 'chat': 150,
            'practice': 15, 'exam': 8, 'walkthrough': 30, 'diagram': 10,
            'srs': 12, 'guide': 4,
        },
        'limits': {
            'course': 25, 'expansion': 50, 'hw_check': 100, 'chat': 250,
            'practice': 25, 'exam': 12, 'walkthrough': 50, 'diagram': 15,
            'srs': 20, 'guide': 8,
        },
    },
}

# 3 Pricing Strategies
PRICING = {
    'A': {'name': 'Standard', 'Student': 29.99, 'Pro': 49.99, 'Family': 79.99},
    'B': {'name': 'Premium',  'Student': 39.99, 'Pro': 69.99, 'Family': 99.99},
    'C': {'name': 'Luxury',   'Student': 49.99, 'Pro': 79.99, 'Family': 129.99},
}

# Plan distribution among subscribers
DIST = {'Student': 0.50, 'Pro': 0.35, 'Family': 0.15}

# Pre-calculate costs
plan_costs_raw = {p: calc_cost(PLANS[p]['usage']) for p in PLANS}
plan_costs_opt = {p: calc_cost(PLANS[p]['usage'], OPT) for p in PLANS}

# ══════════════════════════════════════════════════════════════════
# SHEET 1: USAGE & LIMITS
# ══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Plans and Usage"
ws1.sheet_properties.tabColor = 'FFD600'
NC = 9
for i, w in enumerate([22, 10, 12, 12, 12, 12, 12, 12, 12], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

r = mt(ws1, 1, "NoteSnap Subscription Plans — Usage & Limits", NC)
r = ms(ws1, r, "Pure subscription. No free tier. Every user pays. Tutor replacement pricing.", NC)

# Tutor comparison
r += 1
r = sec(ws1, r, "💰 VALUE: What Parents Pay Now vs NoteSnap", NC, GOLD_BG)
r = hdr(ws1, r, ["", "Private Tutor", "", "Student $29-49", "Pro $49-79", "Family $79-129", "", "", ""])
comp = [
    ("Monthly cost", "$400-800", "", "$29-49", "$49-79", "$79-129"),
    ("Hours/week", "2-3 hours", "", "24/7 unlimited", "24/7 unlimited", "24/7 unlimited"),
    ("Subjects", "1 subject", "", "All subjects", "All subjects", "All subjects"),
    ("Progress tracking", "None", "", "Full analytics", "Full analytics", "Full analytics"),
    ("Students covered", "1 student", "", "1 student", "1 student", "2-5 students"),
    ("SAVINGS vs tutor", "", "", "92-95%", "85-93%", "80-90%"),
]
for i, (label, tutor, _, s, p, f) in enumerate(comp):
    fill = GOLD_BG if i == len(comp)-1 else (WHITE_F if i%2==0 else LBLUE)
    for j, v in enumerate([label, tutor, "", s, p, f], 1):
        c = ws1.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BB if i == len(comp)-1 or j == 1 else BR
        c.alignment = CTR
        if j == 1: c.alignment = LFT
        if i == len(comp)-1 and j >= 4: c.font = GB2
    r += 1

# Usage table
r += 1
r = sec(ws1, r, "📊 Monthly Usage (Avg) & Limits Per Plan", NC)
r = hdr(ws1, r, ["Feature", "Cost", "Student Use", "Student Limit", "Pro Use", "Pro Limit", "Family Use", "Family Limit", ""])

for i, (key, name) in enumerate(FEATURES):
    fill = WHITE_F if i%2==0 else LBLUE
    vals = [
        name, C[key],
        PLANS['Student']['usage'][key], PLANS['Student']['limits'][key],
        PLANS['Pro']['usage'][key], PLANS['Pro']['limits'][key],
        PLANS['Family']['usage'][key], PLANS['Family']['limits'][key],
    ]
    for j, v in enumerate(vals, 1):
        c = ws1.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if j == 1: c.alignment = LFT; c.font = BB
        if j == 2: c.number_format = '$#,##0.000'
        if j in (4,6,8): c.font = Font(name='Calibri', size=10, color='1565C0')  # limits in blue
    r += 1

# Total cost row
r_t = r
for j, v in enumerate(["AI COST PER USER", ""], 1):
    c = ws1.cell(row=r, column=j, value=v)
    c.fill = NAVY; c.font = WB; c.alignment = CTR

cols_cost = [
    ('Student', 3), ('Student', 4), ('Pro', 5), ('Pro', 6), ('Family', 7), ('Family', 8)
]
for plan, col in cols_cost:
    is_limit = col in (4, 6, 8)
    if is_limit:
        cost = calc_cost(PLANS[plan]['limits'], OPT)
        label = f"${cost:.2f} max"
    else:
        cost = plan_costs_opt[plan]
        label = f"${cost:.2f} avg"
    c = ws1.cell(row=r, column=col, value=label)
    c.fill = NAVY; c.font = WB; c.alignment = CTR

r += 2
r = nt(ws1, r, f"💡 Avg AI cost with optimizations: Student=${plan_costs_opt['Student']:.2f} | Pro=${plan_costs_opt['Pro']:.2f} | Family=${plan_costs_opt['Family']:.2f}", NC, LGREEN)
r = nt(ws1, r, f"💡 Raw AI cost (no optimization): Student=${plan_costs_raw['Student']:.2f} | Pro=${plan_costs_raw['Pro']:.2f} | Family=${plan_costs_raw['Family']:.2f}", NC, LORANGE)

# ══════════════════════════════════════════════════════════════════
# SHEET 2: UNIT ECONOMICS — ALL 3 PRICING STRATEGIES
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Unit Economics")
ws2.sheet_properties.tabColor = '2E7D32'
NC = 10
for i, w in enumerate([20, 14, 14, 14, 14, 14, 14, 14, 14, 14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

r = mt(ws2, 1, "Unit Economics — Profit Per Subscriber", NC)
r = ms(ws2, r, "Every user pays. No free tier. With 55% cost optimization.", NC)

for p_key in ['A', 'B', 'C']:
    p = PRICING[p_key]
    r += 1
    r = sec(ws2, r, f"Strategy {p_key}: {p['name']} — Student ${p['Student']} | Pro ${p['Pro']} | Family ${p['Family']}", NC,
           LBLUE if p_key == 'A' else (GOLD_BG if p_key == 'B' else PURPLE))
    
    r = hdr(ws2, r, ["", "Student", "", "Pro", "", "Family", "", "Weighted Avg", "", ""])
    # Sub-headers
    sub = ["Metric", "Raw", "Optimized", "Raw", "Optimized", "Raw", "Optimized", "Raw", "Optimized", ""]
    for j, h in enumerate(sub, 1):
        c = ws2.cell(row=r, column=j, value=h)
        c.fill = DARK; c.font = Font(name='Calibri', bold=True, color='B0BEC5', size=9)
        c.alignment = CTR
    r += 1
    
    rows = [
        "Subscription Price",
        "Stripe Fee (~3.5%)",
        "Net Revenue",
        "AI Cost",
        "═══ PROFIT / USER ═══",
        "Margin %",
    ]
    
    for i, label in enumerate(rows):
        is_total = "PROFIT" in label
        is_margin = "Margin" in label
        fill = NAVY if is_total else (WHITE_F if i%2==0 else LBLUE)
        font = WB if is_total else BR
        
        c = ws2.cell(row=r, column=1, value=label)
        c.fill = fill; c.font = font; c.alignment = LFT
        
        for plan_idx, plan in enumerate(['Student', 'Pro', 'Family']):
            price = p[plan]
            sf = stripe(price)
            net = price - sf
            raw_cost = plan_costs_raw[plan]
            opt_cost = plan_costs_opt[plan]
            
            if label == "Subscription Price":
                raw_v = opt_v = price
            elif label == "Stripe Fee (~3.5%)":
                raw_v = opt_v = sf
            elif label == "Net Revenue":
                raw_v = opt_v = net
            elif label == "AI Cost":
                raw_v = raw_cost; opt_v = opt_cost
            elif "PROFIT" in label:
                raw_v = net - raw_cost; opt_v = net - opt_cost
            elif "Margin" in label:
                raw_v = (net - raw_cost) / net if net > 0 else 0
                opt_v = (net - opt_cost) / net if net > 0 else 0
            
            col_raw = 2 + plan_idx * 2
            col_opt = 3 + plan_idx * 2
            
            for col, v in [(col_raw, raw_v), (col_opt, opt_v)]:
                c = ws2.cell(row=r, column=col, value=v)
                c.fill = fill; c.alignment = CTR
                if is_margin:
                    c.number_format = '0%'
                    c.font = GB if v >= 0 else RB
                elif is_total:
                    c.number_format = '$#,##0.00'
                    c.font = Font(name='Calibri', bold=True, size=12,
                                color='4CAF50' if v >= 0 else 'FF5252')
                else:
                    c.number_format = '$#,##0.00'
                    c.font = font
        
        # Weighted average
        for is_opt, col in [(False, 8), (True, 9)]:
            prices_w = sum(p[plan] * DIST[plan] for plan in DIST)
            sf_w = stripe(prices_w)
            net_w = prices_w - sf_w
            cost_w = sum((plan_costs_opt[plan] if is_opt else plan_costs_raw[plan]) * DIST[plan] for plan in DIST)
            
            if label == "Subscription Price": v = prices_w
            elif "Stripe" in label: v = sf_w
            elif "Net Revenue" in label: v = net_w
            elif "AI Cost" in label: v = cost_w
            elif "PROFIT" in label: v = net_w - cost_w
            elif "Margin" in label: v = (net_w - cost_w) / net_w if net_w > 0 else 0
            
            c = ws2.cell(row=r, column=col, value=v)
            c.fill = fill; c.alignment = CTR
            if is_margin: c.number_format = '0%'; c.font = GB if v >= 0 else RB
            elif is_total:
                c.number_format = '$#,##0.00'
                c.font = Font(name='Calibri', bold=True, size=12, color='4CAF50' if v >= 0 else 'FF5252')
            else: c.number_format = '$#,##0.00'; c.font = font
        
        r += 1
    r += 1


# ══════════════════════════════════════════════════════════════════
# SHEET 3: P&L AT SCALE — ALL 3 STRATEGIES
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("P&L at Scale")
ws3.sheet_properties.tabColor = 'FFD600'
NC = 11
for i, w in enumerate([16, 10, 10, 10, 14, 14, 14, 14, 14, 14, 12], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

r = mt(ws3, 1, "P&L at Scale — All Subscribers Pay (Optimized Costs)", NC)
r = ms(ws3, r, "Distribution: 50% Student | 35% Pro | 15% Family — All figures monthly", NC)

for p_key in ['A', 'B', 'C']:
    p = PRICING[p_key]
    r += 1
    bg = LBLUE if p_key == 'A' else (GOLD_BG if p_key == 'B' else PURPLE)
    r = sec(ws3, r, f"Strategy {p_key}: {p['name']} — ${p['Student']} / ${p['Pro']} / ${p['Family']}", NC, bg)
    r = hdr(ws3, r, [
        "Subscribers", "Student", "Pro", "Family",
        "Total AI Cost", "Stripe Fees", "Infrastructure", "TOTAL COST",
        "GROSS REVENUE", "NET PROFIT", "Margin"
    ])
    
    for idx, total in enumerate(SCALES):
        n_s = int(total * DIST['Student'])
        n_p = int(total * DIST['Pro'])
        n_f = total - n_s - n_p
        
        ai = n_s * plan_costs_opt['Student'] + n_p * plan_costs_opt['Pro'] + n_f * plan_costs_opt['Family']
        
        gross = n_s * p['Student'] + n_p * p['Pro'] + n_f * p['Family']
        sf = n_s * stripe(p['Student']) + n_p * stripe(p['Pro']) + n_f * stripe(p['Family'])
        inf = infra(total)
        tc = ai + sf + inf
        profit = gross - tc
        margin = profit / gross if gross > 0 else 0
        
        fill = WHITE_F if idx%2==0 else bg
        vals = [total, n_s, n_p, n_f, ai, sf, inf, tc, gross, profit, margin]
        
        for j, v in enumerate(vals, 1):
            c = ws3.cell(row=r, column=j, value=v)
            c.fill = fill; c.font = BR; c.alignment = CTR; c.border = THIN
            if j == 1: c.font = BB2
            elif j in (5,6,7): c.number_format = '$#,##0'
            elif j == 8: c.font = RB; c.number_format = '$#,##0'
            elif j == 9: c.font = GOLD_F; c.number_format = '$#,##0'
            elif j == 10:
                c.font = GB2 if v >= 0 else RB
                c.fill = PGREEN if v >= 0 else PRED
                c.number_format = '$#,##0'
            elif j == 11:
                c.number_format = '0%'
                c.font = GB if isinstance(v, float) and v >= 0 else RB
        r += 1
    
    # Annual projection
    for total in [500, 1000, 5000]:
        n_s = int(total * DIST['Student'])
        n_p = int(total * DIST['Pro'])
        n_f = total - n_s - n_p
        ai = n_s * plan_costs_opt['Student'] + n_p * plan_costs_opt['Pro'] + n_f * plan_costs_opt['Family']
        gross = n_s * p['Student'] + n_p * p['Pro'] + n_f * p['Family']
        sf = n_s * stripe(p['Student']) + n_p * stripe(p['Pro']) + n_f * stripe(p['Family'])
        tc = ai + sf + infra(total)
        annual_profit = (gross - tc) * 12
        if total == 1000:
            r = nt(ws3, r, f"📅 Annual at {total:,} subs: ${annual_profit:,.0f}/year profit", NC, PGREEN, GB2)
    r += 1

# ══════════════════════════════════════════════════════════════════
# SHEET 4: SIDE-BY-SIDE WINNER TABLE
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Winner Comparison")
ws4.sheet_properties.tabColor = '1565C0'
NC = 10
for i, w in enumerate([16, 14, 14, 14, 14, 14, 14, 14, 14, 14], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

r = mt(ws4, 1, "Head-to-Head: Which Pricing Wins?", NC)
r = ms(ws4, r, "Monthly profit comparison across all scales — optimized costs", NC)

r += 1
r = hdr(ws4, r, [
    "Subscribers", 
    "A Revenue", "A Cost", "A PROFIT",
    "B Revenue", "B Cost", "B PROFIT",
    "C Revenue", "C Cost", "C PROFIT",
])

for idx, total in enumerate(SCALES):
    fill = WHITE_F if idx%2==0 else LBLUE
    vals = [total]
    best_profit = -999999
    
    for p_key in ['A', 'B', 'C']:
        p = PRICING[p_key]
        n_s = int(total * DIST['Student'])
        n_p = int(total * DIST['Pro'])
        n_f = total - n_s - n_p
        
        ai = n_s * plan_costs_opt['Student'] + n_p * plan_costs_opt['Pro'] + n_f * plan_costs_opt['Family']
        gross = n_s * p['Student'] + n_p * p['Pro'] + n_f * p['Family']
        sf = n_s * stripe(p['Student']) + n_p * stripe(p['Pro']) + n_f * stripe(p['Family'])
        tc = ai + sf + infra(total)
        profit = gross - tc
        
        vals.extend([gross, tc, profit])
        if profit > best_profit: best_profit = profit
    
    for j, v in enumerate(vals, 1):
        c = ws4.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR; c.border = THIN
        if j == 1: c.font = BB2
        elif (j-2) % 3 == 0: c.number_format = '$#,##0'; c.font = GOLD_F  # revenue
        elif (j-3) % 3 == 0: c.number_format = '$#,##0'; c.font = RB       # cost
        elif (j-4) % 3 == 0:  # profit
            c.number_format = '$#,##0'
            c.font = GB2 if v >= 0 else RB
            if isinstance(v, (int,float)) and v == best_profit:
                c.fill = PGREEN
                c.font = Font(name='Calibri', bold=True, size=13, color='1B5E20')
    r += 1

# Annual summary
r += 1
r = sec(ws4, r, "📅 ANNUAL PROFIT PROJECTION", NC, GOLD_BG)
r = hdr(ws4, r, ["Subscribers", "", "", "A: Annual", "", "", "B: Annual", "", "", "C: Annual"])

for total in SCALES:
    vals = [f"{total:,} subs", "", ""]
    for p_key in ['A', 'B', 'C']:
        p = PRICING[p_key]
        n_s = int(total * DIST['Student'])
        n_p = int(total * DIST['Pro'])
        n_f = total - n_s - n_p
        ai = n_s * plan_costs_opt['Student'] + n_p * plan_costs_opt['Pro'] + n_f * plan_costs_opt['Family']
        gross = n_s * p['Student'] + n_p * p['Pro'] + n_f * p['Family']
        sf = n_s * stripe(p['Student']) + n_p * stripe(p['Pro']) + n_f * stripe(p['Family'])
        tc = ai + sf + infra(total)
        annual = (gross - tc) * 12
        vals.extend([annual, "", ""])
    
    fill = WHITE_F if total in (50, 500, 5000) else GOLD_BG
    for j, v in enumerate(vals, 1):
        c = ws4.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if j == 1: c.font = BB2
        if j in (4, 7, 10) and isinstance(v, (int, float)):
            c.number_format = '$#,##0'
            c.font = GB2 if v >= 0 else RB
            if total >= 1000: c.font = GB3
    r += 1

r += 1
r = nt(ws4, r, "🏆 ALL strategies are massively profitable with subscription-only model!", NC, PGREEN, GB2)
r = nt(ws4, r, "💡 Strategy C ($49.99/$79.99/$129.99) has the highest revenue AND profit at every scale", NC, PURPLE)
r = nt(ws4, r, "💡 Strategy A ($29.99/$49.99/$79.99) has the lowest barrier — best for initial launch", NC, LBLUE)
r = nt(ws4, r, "🎯 RECOMMENDATION: Launch with B ($39.99/$69.99/$99.99) — premium but accessible", NC, GOLD_BG, GOLD_BIG)

# ══════════════════════════════════════════════════════════════════
# SHEET 5: RAW vs OPTIMIZED COMPARISON
# ══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Raw vs Optimized")
ws5.sheet_properties.tabColor = 'F44336'
NC = 8
for i, w in enumerate([16, 14, 14, 14, 14, 14, 14, 14], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

r = mt(ws5, 1, "Impact of Cost Optimization — Strategy B", NC)
r = ms(ws5, r, "Same revenue, dramatically different costs. This is why optimization matters.", NC)

p = PRICING['B']

r += 1
r = sec(ws5, r, "WITHOUT Optimization (raw Claude Sonnet costs)", NC, LRED)
r = hdr(ws5, r, ["Subs", "Student", "Pro", "Family", "Total Cost", "Revenue", "PROFIT", "Margin"])

for idx, total in enumerate(SCALES):
    n_s = int(total * DIST['Student']); n_p = int(total * DIST['Pro']); n_f = total - n_s - n_p
    ai = n_s * plan_costs_raw['Student'] + n_p * plan_costs_raw['Pro'] + n_f * plan_costs_raw['Family']
    gross = n_s * p['Student'] + n_p * p['Pro'] + n_f * p['Family']
    sf = n_s * stripe(p['Student']) + n_p * stripe(p['Pro']) + n_f * stripe(p['Family'])
    tc = ai + sf + infra(total)
    profit = gross - tc
    margin = profit / gross if gross > 0 else 0
    fill = WHITE_F if idx%2==0 else LRED
    vals = [total, n_s, n_p, n_f, tc, gross, profit, margin]
    for j, v in enumerate(vals, 1):
        c = ws5.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if j == 1: c.font = BB2
        elif j in (5,6): c.number_format = '$#,##0'
        elif j == 7: c.font = GB if v>=0 else RB; c.number_format = '$#,##0'; c.fill = PGREEN if v>=0 else PRED
        elif j == 8: c.number_format = '0%'
    r += 1

r += 1
r = sec(ws5, r, "WITH 55% Optimization (caching + Haiku + batch API)", NC, LGREEN)
r = hdr(ws5, r, ["Subs", "Student", "Pro", "Family", "Total Cost", "Revenue", "PROFIT", "Margin"])

for idx, total in enumerate(SCALES):
    n_s = int(total * DIST['Student']); n_p = int(total * DIST['Pro']); n_f = total - n_s - n_p
    ai = n_s * plan_costs_opt['Student'] + n_p * plan_costs_opt['Pro'] + n_f * plan_costs_opt['Family']
    gross = n_s * p['Student'] + n_p * p['Pro'] + n_f * p['Family']
    sf = n_s * stripe(p['Student']) + n_p * stripe(p['Pro']) + n_f * stripe(p['Family'])
    tc = ai + sf + infra(total)
    profit = gross - tc
    margin = profit / gross if gross > 0 else 0
    fill = WHITE_F if idx%2==0 else LGREEN
    vals = [total, n_s, n_p, n_f, tc, gross, profit, margin]
    for j, v in enumerate(vals, 1):
        c = ws5.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if j == 1: c.font = BB2
        elif j in (5,6): c.number_format = '$#,##0'
        elif j == 7: c.font = GB2 if v>=0 else RB; c.number_format = '$#,##0'; c.fill = PGREEN if v>=0 else PRED
        elif j == 8: c.number_format = '0%'; c.font = GB
    r += 1

r += 1
# Show the savings
r = sec(ws5, r, "💰 OPTIMIZATION IMPACT AT 1000 SUBSCRIBERS", NC, GOLD_BG)
total = 1000
n_s = int(total * DIST['Student']); n_p = int(total * DIST['Pro']); n_f = total - n_s - n_p
ai_raw = n_s * plan_costs_raw['Student'] + n_p * plan_costs_raw['Pro'] + n_f * plan_costs_raw['Family']
ai_opt = n_s * plan_costs_opt['Student'] + n_p * plan_costs_opt['Pro'] + n_f * plan_costs_opt['Family']
saved = (ai_raw - ai_opt) * 12
r = nt(ws5, r, f"Raw AI cost: ${ai_raw:,.0f}/mo | Optimized: ${ai_opt:,.0f}/mo | Monthly savings: ${ai_raw-ai_opt:,.0f}", NC, LORANGE)
r = nt(ws5, r, f"🔥 Annual savings from optimization: ${saved:,.0f}/year!", NC, PGREEN, GB2)
r = nt(ws5, r, "Optimizations: (1) Prompt caching on system prompts (2) Haiku for chat + classification (3) Batch API for course gen", NC, LBLUE)

# ══════════════════════════════════════════════════════════════════
# SAVE & PRINT
# ══════════════════════════════════════════════════════════════════
out = "/Users/curvalux/NoteSnap/docs/budget/NoteSnap_Subscription_Model.xlsx"
wb.save(out)

print(f"Saved to {out}")
print(f"\n{'='*70}")
print("PLAN COSTS (monthly, per subscriber)")
print(f"{'='*70}")
for plan in ['Student', 'Pro', 'Family']:
    raw = plan_costs_raw[plan]
    opt = plan_costs_opt[plan]
    usage = sum(PLANS[plan]['usage'].values())
    limits = sum(PLANS[plan]['limits'].values())
    print(f"  {plan:8s}: {usage} actions/mo (limit {limits}), raw=${raw:.2f}, opt=${opt:.2f}")

print(f"\n{'='*70}")
print("UNIT ECONOMICS (optimized, per subscriber)")
print(f"{'='*70}")
for p_key in ['A', 'B', 'C']:
    p = PRICING[p_key]
    print(f"\n  Strategy {p_key}: {p['name']}")
    for plan in ['Student', 'Pro', 'Family']:
        price = p[plan]
        sf = stripe(price)
        net = price - sf
        profit = net - plan_costs_opt[plan]
        margin = profit / net * 100
        print(f"    {plan:8s}: ${price:6.2f} - ${sf:.2f} stripe - ${plan_costs_opt[plan]:.2f} AI = ${profit:+.2f} profit ({margin:.0f}% margin)")

print(f"\n{'='*70}")
print("P&L SUMMARY (optimized, monthly)")
print(f"{'='*70}")
for p_key in ['A', 'B', 'C']:
    p = PRICING[p_key]
    print(f"\n  Strategy {p_key}: {p['name']}")
    for total in SCALES:
        n_s = int(total * DIST['Student']); n_p = int(total * DIST['Pro']); n_f = total - n_s - n_p
        ai = n_s * plan_costs_opt['Student'] + n_p * plan_costs_opt['Pro'] + n_f * plan_costs_opt['Family']
        gross = n_s * p['Student'] + n_p * p['Pro'] + n_f * p['Family']
        sf = n_s * stripe(p['Student']) + n_p * stripe(p['Pro']) + n_f * stripe(p['Family'])
        tc = ai + sf + infra(total)
        profit = gross - tc
        annual = profit * 12
        margin = profit / gross * 100
        print(f"    {total:5d} subs: rev=${gross:>8,.0f} cost=${tc:>8,.0f} profit=${profit:>+8,.0f}/mo (${annual:>+10,.0f}/yr) margin={margin:.0f}%")
