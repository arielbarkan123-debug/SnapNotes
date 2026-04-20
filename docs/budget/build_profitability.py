#!/usr/bin/env python3
"""
NoteSnap — Profitability Strategies
"We replace private tutors, not compete with apps"
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ──
NAVY = PatternFill('solid', fgColor='1B2A4A')
DARK = PatternFill('solid', fgColor='2C3E6B')
GOLD_BG = PatternFill('solid', fgColor='F9F3E3')
LBLUE = PatternFill('solid', fgColor='E8EDF5')
LGREEN = PatternFill('solid', fgColor='E8F5E9')
LRED = PatternFill('solid', fgColor='FFEBEE')
LORANGE = PatternFill('solid', fgColor='FFF3E0')
WHITE_F = PatternFill('solid', fgColor='FFFFFF')
PGREEN = PatternFill('solid', fgColor='C8E6C9')
PRED = PatternFill('solid', fgColor='FFCDD2')
GOLD_H = PatternFill('solid', fgColor='FFF8E1')
PURPLE = PatternFill('solid', fgColor='F3E5F5')

WB = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
WT = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
WS = Font(name='Calibri', color='B0BEC5', size=10)
BB = Font(name='Calibri', bold=True, size=11)
BB2 = Font(name='Calibri', bold=True, size=12)
BR = Font(name='Calibri', size=10)
GB = Font(name='Calibri', bold=True, size=11, color='2E7D32')
GB2 = Font(name='Calibri', bold=True, size=13, color='2E7D32')
RB = Font(name='Calibri', bold=True, size=11, color='C62828')
SF = Font(name='Calibri', bold=True, size=11, color='1B2A4A')
GOLD_F = Font(name='Calibri', bold=True, size=11, color='B8860B')
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN = Border(bottom=Side(style='thin', color='E0E0E0'))

# ── Cost Data (verified from codebase) ──
# Costs per action (Claude Sonnet 4.6 multi-call pipeline)
ACTION_COSTS = {
    'course':      0.633,
    'expansion':   0.135,
    'hw_check':    0.237,
    'chat':        0.055,
    'practice':    0.197,
    'exam':        0.137,
    'walkthrough': 0.070,
    'diagram':     0.175,
    'srs':         0.027,
    'guide':       0.264,
}

# Average student monthly usage (generous — student who studies hard)
HEAVY_STUDENT = {
    'course': 8, 'expansion': 15, 'hw_check': 30, 'chat': 80,
    'practice': 10, 'exam': 4, 'walkthrough': 15, 'diagram': 5,
    'srs': 5, 'guide': 2,
}
MEDIUM_STUDENT = {
    'course': 4, 'expansion': 8, 'hw_check': 15, 'chat': 40,
    'practice': 5, 'exam': 2, 'walkthrough': 8, 'diagram': 3,
    'srs': 3, 'guide': 1,
}
LIGHT_STUDENT = {
    'course': 2, 'expansion': 3, 'hw_check': 5, 'chat': 15,
    'practice': 2, 'exam': 0, 'walkthrough': 3, 'diagram': 1,
    'srs': 1, 'guide': 0,
}

def user_cost(usage, optimization=1.0):
    return sum(usage.get(k,0) * v for k,v in ACTION_COSTS.items()) * optimization

def stripe_fee(amount):
    return amount * 0.035 + 0.30

def infra_cost(users):
    supa = 0 if users <= 100 else (25 if users <= 5000 else 599)
    vercel = 20
    resend = 0 if users <= 250 else 20
    domain = 1.25
    return supa + vercel + resend + domain

HEAVY_COST = user_cost(HEAVY_STUDENT)
MEDIUM_COST = user_cost(MEDIUM_STUDENT)
LIGHT_COST = user_cost(LIGHT_STUDENT)

HEAVY_OPT = user_cost(HEAVY_STUDENT, 0.45)
MEDIUM_OPT = user_cost(MEDIUM_STUDENT, 0.45)
LIGHT_OPT = user_cost(LIGHT_STUDENT, 0.45)

SCALES = [50, 100, 500, 1000, 5000]

def mt(ws, row, text, cols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = NAVY; c.font = WT; c.alignment = CTR
    return row + 1

def ms(ws, row, text, cols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = NAVY; c.font = WS; c.alignment = CTR
    return row + 1

def sec(ws, row, text, cols=10, fill=LBLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = SF; c.fill = fill; c.alignment = CTR
    return row + 1

def hdr(ws, row, headers, fill=DARK):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = fill; c.font = WB; c.alignment = CTR
    return row + 1

def note(ws, row, text, cols=10, fill=LGREEN, font=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill; c.font = font or Font(name='Calibri', bold=True, size=10, color='1B2A4A')
    c.alignment = CTR
    return row + 1

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1: THE VALUE FRAME
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Why We Can Charge More"
ws1.sheet_properties.tabColor = 'FFD600'
NC = 6
for i, w in enumerate([30, 16, 16, 16, 16, 16], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

r = mt(ws1, 1, "NoteSnap = Private Tutor Replacement", NC)
r = ms(ws1, r, "We don't compete with apps. We replace $50-80/hour private teachers.", NC)

r += 1
r = sec(ws1, r, "💰 What Parents Currently Pay for Private Tutoring", NC, GOLD_BG)
r = hdr(ws1, r, ["Market", "Hourly Rate", "Hours/Week", "Monthly Cost", "Annual Cost", "NoteSnap Savings"])

tutor_data = [
    ("🇮🇱 Israel", "$27-55", "2-3", "$240-660", "$2,400-6,600", "92-97%"),
    ("🇺🇸 USA", "$40-80", "2-3", "$400-960", "$4,000-9,600", "94-97%"),
    ("🇬🇧 UK", "$35-70", "2-3", "$320-840", "$3,200-8,400", "93-97%"),
    ("🌍 Global Average", "~$50/hr", "2.5 avg", "~$500/mo", "~$5,000/yr", "93-95%"),
]
for i, (market, rate, hrs, monthly, annual, save) in enumerate(tutor_data):
    fill = WHITE_F if i % 2 == 0 else GOLD_BG
    for j, v in enumerate([market, rate, hrs, monthly, annual, save], 1):
        c = ws1.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BB if j in (4,6) else BR; c.alignment = CTR
        if j == 1: c.alignment = LFT
        if j == 6: c.font = GB2
    r += 1

r += 1
r = note(ws1, r, "🎯 Even at $49.99/month, NoteSnap saves parents 90%+ vs private tutors!", NC, PGREEN, GB2)
r = note(ws1, r, "🎯 NoteSnap is available 24/7, covers ALL subjects, gives instant feedback, and tracks progress", NC, PGREEN)

# Student cost table
r += 1
r = sec(ws1, r, "📊 What Each Student Costs Us in AI (Monthly)", NC)
r = hdr(ws1, r, ["Usage Level", "Actions/Month", "AI Cost (raw)", "AI Cost (optimized)", "Description", ""])

profiles = [
    ("🟢 Light", sum(LIGHT_STUDENT.values()), LIGHT_COST, LIGHT_OPT, "Casual user, HW help only"),
    ("🟡 Medium", sum(MEDIUM_STUDENT.values()), MEDIUM_COST, MEDIUM_OPT, "Regular student, most features"),
    ("🔴 Heavy", sum(HEAVY_STUDENT.values()), HEAVY_COST, HEAVY_OPT, "Power user, exam season mode"),
]
for i, (name, actions, raw, opt, desc) in enumerate(profiles):
    fill = WHITE_F if i % 2 == 0 else LBLUE
    for j, v in enumerate([name, actions, raw, opt, desc], 1):
        c = ws1.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BB if j in (3,4) else BR; c.alignment = CTR
        if j == 1: c.alignment = LFT; c.font = BB
        if j in (3,4) and isinstance(v, float): c.number_format = '$#,##0.00'; c.font = RB
    r += 1

r += 1
r = note(ws1, r, f"Raw costs: Light=${LIGHT_COST:.2f} | Medium=${MEDIUM_COST:.2f} | Heavy=${HEAVY_COST:.2f}", NC, LORANGE)
r = note(ws1, r, f"With optimization (caching+Haiku+batch): Light=${LIGHT_OPT:.2f} | Medium=${MEDIUM_OPT:.2f} | Heavy=${HEAVY_OPT:.2f}", NC, LGREEN)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 2: 8 PROFITABILITY STRATEGIES
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("8 Ways to Profit")
ws2.sheet_properties.tabColor = '2E7D32'
NC = 8
for i, w in enumerate([24, 14, 14, 14, 14, 14, 14, 14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

r = mt(ws2, 1, "8 Profitability Strategies — Pick Your Path", NC)
r = ms(ws2, r, "Each strategy compared at 500 active users, 55% cost optimization applied", NC)

strategies = []

# ── STRATEGY 1: Premium Flat Rate ──
r += 1
r = sec(ws2, r, "① PREMIUM FLAT RATE — One price, unlimited use", NC, GOLD_BG)
r = hdr(ws2, r, ["Price/Month", "Conv %", "Paying Users", "Free Users", "Total Cost", "Net Revenue", "PROFIT", "Margin"])

for price in [29.99, 39.99, 49.99, 59.99]:
    conv = {29.99: 0.10, 39.99: 0.08, 49.99: 0.06, 59.99: 0.05}[price]
    n_paying = int(500 * conv)
    n_free = 500 - n_paying
    # Paying users: 40% medium, 60% heavy (premium attracts serious students)
    ai_pay = n_paying * (0.4 * MEDIUM_OPT + 0.6 * HEAVY_OPT)
    ai_free = n_free * LIGHT_OPT * 0.3  # minimal free tier
    tc = ai_pay + ai_free + infra_cost(500)
    sf = n_paying * stripe_fee(price)
    nr = n_paying * price - sf
    profit = nr - tc
    margin = profit / nr if nr > 0 else -1
    
    fill = WHITE_F if price in (29.99, 49.99) else LBLUE
    vals = [f"${price}", f"{conv*100:.0f}%", n_paying, n_free, tc, nr, profit, margin]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if j == 5: c.number_format = '$#,##0'; c.font = RB
        if j == 6: c.number_format = '$#,##0'; c.font = GB
        if j == 7: c.font = GB if isinstance(v,(int,float)) and v>=0 else RB; c.number_format = '$#,##0'
        if j == 7: c.fill = PGREEN if isinstance(v,(int,float)) and v>=0 else PRED
        if j == 8 and isinstance(v, float): c.number_format = '0%'
    
    strategies.append(('Premium Flat', price, conv, profit, margin if isinstance(margin, float) else 0))
    r += 1

r += 1

# ── STRATEGY 2: Credit/Token System ──
r = sec(ws2, r, "② CREDIT PACKS — Pay per use, every action profitable", NC, PURPLE)

# Credits pricing: each credit = ~$1 worth of usage
credit_packs = [
    ("Starter", 30, 14.99),
    ("Standard", 75, 29.99),
    ("Pro", 150, 49.99),
    ("Mega", 300, 79.99),
]

# Credit costs per action
credit_actions = [
    ("Course Generation", 'course', 3),
    ("Homework Check", 'hw_check', 1),
    ("Chat Message", 'chat', 1),
    ("Practice Session", 'practice', 1),
    ("Exam Generation", 'exam', 2),
    ("Walkthrough", 'walkthrough', 1),
    ("Diagram", 'diagram', 1),
    ("Lesson Expansion", 'expansion', 1),
    ("Study Guide", 'guide', 2),
    ("SRS Batch", 'srs', 1),
]

r = hdr(ws2, r, ["Credit Pack", "Credits", "Price", "$/Credit", "Our Cost/Cr", "Profit/Cr", "Margin", ""])

for name, credits, price in credit_packs:
    cost_per_credit = price / credits
    # Average action cost (weighted by typical usage)
    avg_action_cost = sum(ACTION_COSTS[k] * MEDIUM_STUDENT.get(k,0) for k in ACTION_COSTS) / sum(MEDIUM_STUDENT.values()) * 0.45
    profit_per = cost_per_credit - avg_action_cost
    margin = profit_per / cost_per_credit
    
    fill = WHITE_F if name in ('Starter', 'Pro') else PURPLE
    vals = [f"🎫 {name}", credits, price, cost_per_credit, avg_action_cost, profit_per, margin]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if j >= 3 and isinstance(v, float): c.number_format = '$#,##0.00'
        if j == 6: c.font = GB; c.number_format = '$#,##0.00'
        if j == 7: c.number_format = '0%'; c.font = GB
    r += 1

# P&L at 500 users
r += 1
r = note(ws2, r, "📊 Credits P&L: At 500 users, avg 50 credits/mo purchased, Standard pack ($29.99)", NC, LBLUE)
conv_c = 0.12  # credits convert better (lower commitment)
n_buy = int(500 * conv_c)
avg_rev = 29.99  # average pack
ai_c = n_buy * MEDIUM_OPT  # medium usage for credit buyers
ai_free_c = (500 - n_buy) * LIGHT_OPT * 0.3
tc_c = ai_c + ai_free_c + infra_cost(500)
nr_c = n_buy * avg_rev - n_buy * stripe_fee(avg_rev)
profit_c = nr_c - tc_c
strategies.append(('Credit Packs', 29.99, conv_c, profit_c, profit_c/nr_c if nr_c > 0 else 0))

r = hdr(ws2, r, ["", "Buyers", "Free", "Total Cost", "Net Revenue", "PROFIT", "Margin", ""])
vals = ["@500 users, 12% buy", n_buy, 500-n_buy, tc_c, nr_c, profit_c, profit_c/nr_c]
for j, v in enumerate(vals, 1):
    c = ws2.cell(row=r, column=j, value=v)
    c.fill = LGREEN if isinstance(v,(int,float)) and j==6 and v>0 else WHITE_F
    c.font = BR; c.alignment = CTR
    if j in (4,5,6) and isinstance(v, float): c.number_format = '$#,##0'
    if j == 6: c.font = GB if v >= 0 else RB
    if j == 7 and isinstance(v, float): c.number_format = '0%'
r += 2

# ── STRATEGY 3: Subject Packs ──
r = sec(ws2, r, "③ SUBJECT PACKS — Pay per subject area", NC, LORANGE)
r = hdr(ws2, r, ["Plan", "Subjects", "Price", "Typical Use", "AI Cost", "Stripe", "Profit/User", "Margin"])

subject_plans = [
    ("1 Subject", 1, 14.99, 0.35),   # 35% of full usage
    ("3 Subjects", 3, 29.99, 0.70),
    ("Unlimited", 99, 44.99, 1.0),
]
for i, (name, subj, price, usage_mult) in enumerate(subject_plans):
    ai = MEDIUM_OPT * usage_mult
    sf = stripe_fee(price)
    profit = price - sf - ai
    margin = profit / (price - sf)
    fill = WHITE_F if i % 2 == 0 else LORANGE
    vals = [name, subj if subj < 99 else "∞", price, f"{usage_mult*100:.0f}%", ai, sf, profit, margin]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if isinstance(v, float) and j in (3,5,6): c.number_format = '$#,##0.00'
        if j == 7: c.font = GB if isinstance(v,(int,float)) and v>=0 else RB; c.number_format = '$#,##0.00'
        if j == 8 and isinstance(v, float): c.number_format = '0%'
    r += 1

conv_subj = 0.09
n_pay_s = int(500 * conv_subj)
# Mix: 50% buy 1 subj, 35% buy 3, 15% unlimited
ai_s = n_pay_s * (0.50 * MEDIUM_OPT*0.35 + 0.35 * MEDIUM_OPT*0.70 + 0.15 * MEDIUM_OPT)
ai_free_s = (500-n_pay_s) * LIGHT_OPT * 0.3
tc_s = ai_s + ai_free_s + infra_cost(500)
avg_price_s = 0.50*14.99 + 0.35*29.99 + 0.15*44.99
nr_s = n_pay_s * avg_price_s - n_pay_s * stripe_fee(avg_price_s)
profit_s = nr_s - tc_s
strategies.append(('Subject Packs', avg_price_s, conv_subj, profit_s, profit_s/nr_s if nr_s > 0 else 0))
r += 1

# ── STRATEGY 4: Family Plans ──
r = sec(ws2, r, "④ FAMILY PLANS — Higher ARPU per household", NC, LGREEN)
r = hdr(ws2, r, ["Plan", "Students", "Price", "Price/Student", "AI Cost/Fam", "Stripe", "Profit/Fam", "Margin"])

family_plans = [
    ("Solo Student", 1, 34.99),
    ("Family (3)", 3, 54.99),
    ("Family+ (5)", 5, 74.99),
]
for i, (name, students, price) in enumerate(family_plans):
    # Family members: 1 heavy, rest medium
    ai = (1 * HEAVY_OPT + max(0, students-1) * MEDIUM_OPT)
    sf = stripe_fee(price)
    profit = price - sf - ai
    margin = profit / (price - sf)
    fill = WHITE_F if i % 2 == 0 else LGREEN
    vals = [name, students, price, price/students, ai, sf, profit, margin]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if isinstance(v, float): c.number_format = '$#,##0.00'
        if j == 7: c.font = GB if v>=0 else RB
        if j == 8: c.number_format = '0%'
    r += 1
r += 1

# ── STRATEGY 5: Hybrid Base + Credits ──
r = sec(ws2, r, "⑤ HYBRID: Low Base + AI Credit Topups", NC, LBLUE)
r = note(ws2, r, "Base subscription gives platform access + limited AI. Buy credit packs when you need more.", NC, LBLUE)
r = hdr(ws2, r, ["Plan", "Base Price", "Included Cr", "Extra 50 Cr", "Avg Revenue", "AI Cost", "Profit/User", "Margin"])

hybrid_plans = [
    ("Starter", 9.99, 20, 14.99),
    ("Plus", 19.99, 50, 14.99),
    ("Premium", 34.99, 100, 14.99),
]
for i, (name, base, included_cr, extra_price) in enumerate(hybrid_plans):
    # Assume avg user buys 1 extra pack per month
    avg_rev = base + extra_price * 0.4  # 40% buy extras
    avg_credits = included_cr + 50 * 0.4
    ai = MEDIUM_OPT  # medium usage
    sf = stripe_fee(avg_rev)
    profit = avg_rev - sf - ai
    margin = profit / (avg_rev - sf)
    fill = WHITE_F if i % 2 == 0 else LBLUE
    vals = [name, base, included_cr, extra_price, avg_rev, ai, profit, margin]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if isinstance(v, float): c.number_format = '$#,##0.00'
        if j == 7: c.font = GB if v>=0 else RB
        if j == 8: c.number_format = '0%'
    r += 1
r += 1

# ── STRATEGY 6: Exam Season Pass ──
r = sec(ws2, r, "⑥ EXAM SEASON PASS — Urgency-based pricing", NC, PRED)
r = hdr(ws2, r, ["Product", "Price", "Duration", "Typical Use", "AI Cost", "Profit", "Margin", "Notes"])

exam_products = [
    ("🔥 Exam Week", 19.99, "7 days", HEAVY_OPT * 0.35, "One exam burst"),
    ("📚 Exam Month", 39.99, "30 days", HEAVY_OPT, "Full exam period"),
    ("🎓 Semester", 99.99, "4 months", MEDIUM_OPT * 4, "Whole semester"),
    ("👑 Annual", 249.99, "12 months", MEDIUM_OPT * 12, "Best value"),
]
for i, (name, price, duration, ai, notes) in enumerate(exam_products):
    sf = stripe_fee(price)
    profit = price - sf - ai
    margin = profit / (price - sf)
    fill = WHITE_F if i % 2 == 0 else PRED
    vals = [name, price, duration, ai, ai, profit, margin, notes]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if isinstance(v, float) and j in (2,4,5,6): c.number_format = '$#,##0.00'
        if j == 6: c.font = GB if v>=0 else RB
        if j == 7 and isinstance(v, float): c.number_format = '0%'
    r += 1
r += 1

# ── STRATEGY 7: B2B School Licensing ──
r = sec(ws2, r, "⑦ B2B SCHOOL LICENSING — Schools pay, students use free", NC, GOLD_BG)
r = hdr(ws2, r, ["School Size", "Students", "Price/Student", "Monthly Rev", "AI Cost", "Profit", "Margin", ""])

school_plans = [
    ("Small School", 50, 12.99),
    ("Medium School", 200, 9.99),
    ("Large School", 500, 7.99),
    ("District", 2000, 5.99),
]
for i, (name, students, pps) in enumerate(school_plans):
    rev = students * pps
    # Schools: mix of usage (some students barely use it)
    ai = students * (0.3 * HEAVY_OPT + 0.4 * MEDIUM_OPT + 0.3 * LIGHT_OPT)
    sf = stripe_fee(rev)
    profit = rev - sf - ai - infra_cost(students)
    margin = profit / (rev - sf)
    fill = WHITE_F if i % 2 == 0 else GOLD_BG
    vals = [name, students, pps, rev, ai, profit, margin]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if isinstance(v, float): c.number_format = '$#,##0.00' if v < 100 else '$#,##0'
        if j == 6: c.font = GB if v>=0 else RB; c.number_format = '$#,##0'
        if j == 7 and isinstance(v, float): c.number_format = '0%'
    r += 1
r += 1

# ── STRATEGY 8: Cost Reduction (make ANY pricing work) ──
r = sec(ws2, r, "⑧ COST REDUCTION — Make any pricing strategy work", NC, LGREEN)
r = hdr(ws2, r, ["Optimization", "Savings", "Medium User Cost", "From", "Impact", "", "", ""])

optimizations = [
    ("None (current)", "0%", MEDIUM_COST, "$0", "Baseline"),
    ("+ Prompt Caching", "30-40%", MEDIUM_COST * 0.65, f"-${MEDIUM_COST*0.35:.2f}", "Cache system prompts"),
    ("+ Haiku for chat/QA", "50%", MEDIUM_COST * 0.50, f"-${MEDIUM_COST*0.50:.2f}", "80% cheaper for simple tasks"),
    ("+ Batch API", "55%", MEDIUM_COST * 0.45, f"-${MEDIUM_COST*0.55:.2f}", "Async generation 50% off"),
    ("+ Anthropic Startup $", "75%", MEDIUM_COST * 0.25, f"-${MEDIUM_COST*0.75:.2f}", "$25-100K free credits"),
]
for i, (name, savings, cost, impact, desc) in enumerate(optimizations):
    fill = WHITE_F if i % 2 == 0 else LGREEN
    if i == len(optimizations)-1: fill = PGREEN
    vals = [name, savings, cost, impact, desc]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BR; c.alignment = CTR
        if j == 1: c.alignment = LFT; c.font = BB
        if j == 3 and isinstance(v, float): c.number_format = '$#,##0.00'; c.font = GB if i > 0 else RB
    r += 1


# ══════════════════════════════════════════════════════════════════════════
# SHEET 3: FULL P&L — TOP 3 STRATEGIES
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Full P&L Comparison")
ws3.sheet_properties.tabColor = 'FFD600'
NC = 10
for i, w in enumerate([18, 10, 10, 10, 14, 14, 14, 14, 14, 12], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

r = mt(ws3, 1, "Full P&L — Best 3 Strategies at Every Scale", NC)
r = ms(ws3, r, "All with 'Light' free tier ($0.97/user) + 55% cost optimization", NC)

# Strategy configs for P&L
top_strategies = [
    {
        'name': '① Premium $39.99/mo',
        'fill': GOLD_BG,
        'conv': 0.08,
        'price': 39.99,
        'mix': (0.0, 0.4, 0.6),  # (light, medium, heavy) of paying users
    },
    {
        'name': '② Credits (avg $29.99)',
        'fill': PURPLE,
        'conv': 0.12,
        'price': 29.99,
        'mix': (0.1, 0.6, 0.3),
    },
    {
        'name': '⑤ Hybrid $19.99 + topup',
        'fill': LBLUE,
        'conv': 0.15,
        'price': 25.99,  # base + avg topup
        'mix': (0.2, 0.5, 0.3),
    },
]

for strat in top_strategies:
    r += 1
    r = sec(ws3, r, f"{strat['name']} — Conversion: {strat['conv']*100:.0f}%", NC, strat['fill'])
    r = hdr(ws3, r, ["Users", "Free", "Paying", "Free Cost", "Paying Cost", "Infra", "TOTAL COST", "NET REVENUE", "PROFIT", "Margin"])
    
    for idx, total in enumerate(SCALES):
        n_pay = max(1, int(total * strat['conv']))
        n_free = total - n_pay
        
        free_ai = n_free * LIGHT_OPT * 0.3  # minimal free tier
        
        l_mix, m_mix, h_mix = strat['mix']
        pay_ai = n_pay * (l_mix * LIGHT_OPT + m_mix * MEDIUM_OPT + h_mix * HEAVY_OPT)
        
        infra = infra_cost(total)
        tc = free_ai + pay_ai + infra
        
        gross = n_pay * strat['price']
        sf = n_pay * stripe_fee(strat['price'])
        nr = gross - sf
        profit = nr - tc
        margin = profit / nr if nr > 0 else -1
        
        fill = WHITE_F if idx % 2 == 0 else strat['fill']
        vals = [total, n_free, n_pay, free_ai, pay_ai, infra, tc, nr, profit, margin]
        
        for j, v in enumerate(vals, 1):
            c = ws3.cell(row=r, column=j, value=v)
            c.fill = fill; c.font = BR; c.alignment = CTR; c.border = THIN
            if j == 1: c.font = BB2
            elif j == 7: c.font = RB; c.number_format = '$#,##0'
            elif j == 8: c.font = GB; c.number_format = '$#,##0'
            elif j == 9:
                c.font = GB if v >= 0 else RB
                c.fill = PGREEN if v >= 0 else PRED
                c.number_format = '$#,##0'
            elif j == 10: c.number_format = '0%'; c.font = GB if isinstance(v,float) and v >= 0 else RB
            elif j in (4,5,6): c.number_format = '$#,##0'
        r += 1

# Grand summary
r += 2
r = sec(ws3, r, "🏆 ANNUAL PROJECTION AT 1000 USERS", NC, GOLD_BG)
r = hdr(ws3, r, ["Strategy", "Monthly Rev", "Monthly Cost", "Monthly Profit", "Annual Profit", "Break-Even", "", "", "", ""])

for strat in top_strategies:
    total = 1000
    n_pay = max(1, int(total * strat['conv']))
    n_free = total - n_pay
    free_ai = n_free * LIGHT_OPT * 0.3
    l,m,h = strat['mix']
    pay_ai = n_pay * (l * LIGHT_OPT + m * MEDIUM_OPT + h * HEAVY_OPT)
    tc = free_ai + pay_ai + infra_cost(total)
    gross = n_pay * strat['price']
    sf = n_pay * stripe_fee(strat['price'])
    nr = gross - sf
    profit = nr - tc
    annual = profit * 12
    
    # Break even
    be = "N/A"
    for t in range(10, 10001, 10):
        np2 = max(1, int(t * strat['conv']))
        nf2 = t - np2
        tc2 = nf2 * LIGHT_OPT * 0.3 + np2 * (l*LIGHT_OPT+m*MEDIUM_OPT+h*HEAVY_OPT) + infra_cost(t)
        nr2 = np2 * strat['price'] - np2 * stripe_fee(strat['price'])
        if nr2 >= tc2:
            be = f"{t} users"
            break
    
    fill = PGREEN if profit > 0 else PRED
    vals = [strat['name'], nr, tc, profit, annual, be]
    for j, v in enumerate(vals, 1):
        c = ws3.cell(row=r, column=j, value=v)
        c.fill = fill; c.font = BB; c.alignment = CTR
        if isinstance(v, float): c.number_format = '$#,##0'
        if j == 4: c.font = GB2 if v >= 0 else RB
        if j == 5: c.font = GB2 if isinstance(v, float) and v >= 0 else RB; c.number_format = '$#,##0'
    r += 1

r += 1
r = note(ws3, r, "💡 ALL 3 STRATEGIES ARE PROFITABLE at every scale with cost optimization!", NC, PGREEN, GB2)
r = note(ws3, r, "💡 Premium Flat Rate gives highest margins | Credits give highest conversion | Hybrid balances both", NC, LGREEN)
r = note(ws3, r, "🎯 RECOMMENDED: Start with Premium $39.99 (simple, high margin) → add Credits later for growth", NC, GOLD_BG, GOLD_F)


# ══════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════
out = "/Users/curvalux/NoteSnap/docs/budget/NoteSnap_Profitability_Strategies.xlsx"
wb.save(out)
print(f"\nSaved to {out}")
print(f"\n{'='*70}")
print(f"STUDENT AI COSTS (per month)")
print(f"{'='*70}")
print(f"  Heavy (raw):  ${HEAVY_COST:.2f}")
print(f"  Medium (raw): ${MEDIUM_COST:.2f}")
print(f"  Light (raw):  ${LIGHT_COST:.2f}")
print(f"  Heavy (opt):  ${HEAVY_OPT:.2f}")
print(f"  Medium (opt): ${MEDIUM_OPT:.2f}")
print(f"  Light (opt):  ${LIGHT_OPT:.2f}")
print(f"\n{'='*70}")
print(f"TOP 3 STRATEGIES @ 1000 USERS (optimized costs)")
print(f"{'='*70}")
for strat in top_strategies:
    total = 1000
    n_pay = max(1, int(total * strat['conv']))
    n_free = total - n_pay
    l,m,h = strat['mix']
    pay_ai = n_pay * (l*LIGHT_OPT+m*MEDIUM_OPT+h*HEAVY_OPT)
    free_ai = n_free * LIGHT_OPT * 0.3
    tc = free_ai + pay_ai + infra_cost(total)
    nr = n_pay * strat['price'] - n_pay * stripe_fee(strat['price'])
    profit = nr - tc
    print(f"  {strat['name']}: {n_pay} paying, rev=${nr:.0f}, cost=${tc:.0f}, profit=${profit:+.0f}/mo (${profit*12:+,.0f}/yr)")
