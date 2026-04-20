#!/usr/bin/env python3
"""NoteSnap Clean P&L Model — Pure Costs vs Revenue"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = Workbook()

# ── Styles ──
NAVY = PatternFill('solid', fgColor='1B2A4A')
DARK_HEADER = PatternFill('solid', fgColor='2C3E6B')
LIGHT_BLUE = PatternFill('solid', fgColor='E8EDF5')
LIGHT_GREEN = PatternFill('solid', fgColor='E8F5E9')
LIGHT_RED = PatternFill('solid', fgColor='FFEBEE')
LIGHT_YELLOW = PatternFill('solid', fgColor='FFFFF0')
WHITE = PatternFill('solid', fgColor='FFFFFF')
PROFIT_GREEN = PatternFill('solid', fgColor='C8E6C9')
LOSS_RED = PatternFill('solid', fgColor='FFCDD2')

W_BOLD = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
W_TITLE = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
W_SUB = Font(name='Calibri', color='B0BEC5', size=10)
B_BOLD = Font(name='Calibri', bold=True, size=11)
B_BOLD_BIG = Font(name='Calibri', bold=True, size=12)
B_REG = Font(name='Calibri', size=10)
B_SMALL = Font(name='Calibri', size=9, color='666666')
GREEN_BOLD = Font(name='Calibri', bold=True, size=11, color='2E7D32')
RED_BOLD = Font(name='Calibri', bold=True, size=11, color='C62828')
BLUE_NUM = Font(name='Calibri', size=10, color='1565C0')
SECTION_FONT = Font(name='Calibri', bold=True, size=11, color='1B2A4A')

THIN = Border(bottom=Side(style='thin', color='E0E0E0'))
MED = Border(bottom=Side(style='medium', color='1B2A4A'))

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

USD = '#,##0.00'
USD_INT = '#,##0'
PCT = '0%'

def style_row(ws, row, cols, fill=None, font=None, align=None, border=None, fmt=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        if fill: cell.fill = fill
        if font: cell.font = font
        if align: cell.alignment = align
        if border: cell.border = border
        if fmt and c > 1: cell.number_format = fmt

def write_row(ws, row, values, fill=None, font=None, align=LEFT, fmt=None):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        if fill: cell.fill = fill
        if font: cell.font = font
        cell.alignment = align if i == 1 else CENTER
        if fmt and i > 1 and isinstance(v, (int, float)):
            cell.number_format = fmt

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1: How Students ACTUALLY Use NoteSnap
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Student Usage"
ws1.sheet_properties.tabColor = '1B2A4A'

# Column widths
for col, w in [(1,30),(2,14),(3,14),(4,14),(5,14),(6,16)]:
    ws1.column_dimensions[get_column_letter(col)].width = w

r = 1
# Title
ws1.merge_cells('A1:F1')
c = ws1.cell(row=1, column=1, value="How Students ACTUALLY Use NoteSnap")
c.fill = NAVY; c.font = W_TITLE; c.alignment = CENTER
r = 2
ws1.merge_cells('A2:F2')
c = ws1.cell(row=2, column=1, value="Realistic monthly usage based on student behavior patterns")
c.fill = NAVY; c.font = W_SUB; c.alignment = CENTER

# Section: Per-action costs
r = 4
ws1.merge_cells(f'A{r}:F{r}')
c = ws1.cell(row=r, column=1, value="💰 Cost Per AI Action (Claude Sonnet 4.6)")
c.font = SECTION_FONT; c.fill = LIGHT_BLUE

r = 5
headers = ["Action", "Input Tokens", "Output Tokens", "Input Cost", "Output Cost", "Total Cost"]
write_row(ws1, r, headers, fill=DARK_HEADER, font=W_BOLD, align=CENTER)

actions = [
    ("📚 Course Generation", 8000, 16000),
    ("📖 Lesson Expansion", 4000, 8000),
    ("✏️ Homework Check", 3000, 4000),
    ("💬 Chat Message", 4000, 1500),
    ("🎯 Practice Session", 3000, 4000),
    ("📝 Exam Generation", 4000, 8000),
    ("🔍 Walkthrough Step", 3000, 4000),
    ("📊 Diagram (Claude+Recraft)", 3000, 3000),
]

for i, (name, inp, out) in enumerate(actions):
    r += 1
    inp_cost = inp / 1_000_000 * 3
    out_cost = out / 1_000_000 * 15
    if "Diagram" in name:
        out_cost += 0.04  # Recraft API
    total = inp_cost + out_cost
    fill = WHITE if i % 2 == 0 else LIGHT_BLUE
    values = [name, f"{inp:,}", f"{out:,}", inp_cost, out_cost, total]
    for j, v in enumerate(values, 1):
        cell = ws1.cell(row=r, column=j, value=v)
        cell.fill = fill
        cell.font = B_REG
        cell.alignment = LEFT if j == 1 else CENTER
        if j >= 4:
            cell.number_format = '$#,##0.000'
    ws1.cell(row=r, column=1).font = B_BOLD

# Section: Usage per student type
r += 2
ws1.merge_cells(f'A{r}:F{r}')
c = ws1.cell(row=r, column=1, value="👨‍🎓 Monthly Usage Per Student Type")
c.font = SECTION_FONT; c.fill = LIGHT_BLUE

r += 1
headers2 = ["Action", "Cost/Action", "Free Student", "Basic ($9.99)", "Pro ($19.99)", ""]
write_row(ws1, r, headers2, fill=DARK_HEADER, font=W_BOLD, align=CENTER)

# Usage data: (action_name, cost_per, free_count, basic_count, pro_count)
usage = [
    ("Course Generation", 0.264, 1, 4, 8),
    ("Homework Check", 0.069, 5, 20, 40),
    ("Chat Messages", 0.035, 8, 30, 70),
    ("Practice Sessions", 0.069, 2, 4, 10),
    ("Exam Generation", 0.132, 0, 1, 3),
    ("Walkthrough Steps", 0.069, 0, 5, 12),
    ("Diagram Generation", 0.094, 1, 3, 8),
]

free_total = 0
basic_total = 0
pro_total = 0

for i, (name, cost, free, basic, pro) in enumerate(usage):
    r += 1
    fill = WHITE if i % 2 == 0 else LIGHT_YELLOW
    free_cost = free * cost
    basic_cost = basic * cost
    pro_cost = pro * cost
    free_total += free_cost
    basic_total += basic_cost
    pro_total += pro_cost
    
    for j, v in enumerate([name, cost, free, basic, pro], 1):
        cell = ws1.cell(row=r, column=j, value=v)
        cell.fill = fill
        cell.font = B_REG
        cell.alignment = LEFT if j == 1 else CENTER
        if j == 2: cell.number_format = '$#,##0.000'

# Totals row
r += 1
for j, v in enumerate(["TOTAL AI COST / MONTH", "", f"${free_total:.2f}", f"${basic_total:.2f}", f"${pro_total:.2f}"], 1):
    cell = ws1.cell(row=r, column=j, value=v)
    cell.fill = NAVY; cell.font = W_BOLD; cell.alignment = CENTER

# Key insight
r += 2
ws1.merge_cells(f'A{r}:F{r}')
c = ws1.cell(row=r, column=1, value=f"⚡ Key: Free user costs ${free_total:.2f}/mo to serve | Basic user nets ${9.99-basic_total:.2f}/mo profit | Pro user nets ${19.99-pro_total:.2f}/mo profit")
c.font = Font(name='Calibri', bold=True, size=11, color='1B2A4A')
c.fill = LIGHT_GREEN
c.alignment = CENTER

# ══════════════════════════════════════════════════════════════════════════
# SHEET 2: THE MAIN P&L TABLE
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("P&L by Scale")
ws2.sheet_properties.tabColor = '2E7D32'

# Columns
col_widths = [22, 10, 10, 10, 13, 10, 10, 10, 10, 13, 13, 10, 13, 13, 12]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Title
r = 1
ws2.merge_cells('A1:O1')
c = ws2.cell(row=1, column=1, value="NoteSnap P&L — Pure Costs vs Revenue")
c.fill = NAVY; c.font = W_TITLE; c.alignment = CENTER
r = 2
ws2.merge_cells('A2:O2')
c = ws2.cell(row=2, column=1, value="Conversion: 60% Free | 25% Basic ($9.99) | 15% Pro ($19.99) — All monthly figures")
c.fill = NAVY; c.font = W_SUB; c.alignment = CENTER

# Headers
r = 4
# Group headers
groups = [
    ("", 1, 1),
    ("👥 USERS", 2, 4),
    ("💸 COSTS", 5, 9),
    ("💰 REVENUE", 10, 13),
    ("📊 RESULT", 14, 15),
]
for label, start, end in groups:
    if start == end:
        continue
    ws2.merge_cells(start_row=r, start_column=start, end_row=r, end_column=end)
    c = ws2.cell(row=r, column=start, value=label)
    c.fill = DARK_HEADER; c.font = W_BOLD; c.alignment = CENTER
    for cc in range(start, end+1):
        ws2.cell(row=r, column=cc).fill = DARK_HEADER

r = 5
headers = [
    "Scale",
    "Free", "Basic", "Pro",
    "AI API", "Supabase", "Vercel", "Stripe Fees", "Total Cost",
    "Basic Rev", "Pro Rev", "Total Rev", "- Stripe", "Net Revenue",
    "PROFIT", 
]
# Oops, that's 15 but I only want clean. Let me simplify.

# Actually let me redo this more cleanly
headers = [
    "Total Users",
    "Free", "Basic", "Pro",
    "AI API", "Supabase", "Vercel", "Other", "TOTAL COST",
    "Basic ×$9.99", "Pro ×$19.99", "Gross Rev", "Stripe 2.9%", "Net Revenue",
    "PROFIT/LOSS",
]

for j, h in enumerate(headers, 1):
    cell = ws2.cell(row=r, column=j, value=h)
    cell.fill = DARK_HEADER; cell.font = W_BOLD; cell.alignment = CENTER
    cell.border = MED

# Data rows
scales = [50, 100, 250, 500, 1000, 2000, 5000]

for idx, total in enumerate(scales):
    r += 1
    free = int(total * 0.60)
    basic = int(total * 0.25)
    pro = total - free - basic  # remainder to pro
    
    # Costs
    ai_cost = free * free_total + basic * basic_total + pro * pro_total
    
    # Supabase
    if total <= 200:
        supa = 0  # free tier
    elif total <= 5000:
        supa = 25  # pro
    else:
        supa = 599  # team
    
    vercel = 20  # pro plan
    
    # Other (Resend + Domain)
    resend = 0 if total <= 300 else 20
    domain = 1.25  # $15/year
    other = resend + domain
    
    total_cost = ai_cost + supa + vercel + other
    
    # Revenue
    basic_rev = basic * 9.99
    pro_rev = pro * 19.99
    gross_rev = basic_rev + pro_rev
    stripe_fee = gross_rev * 0.029 + (basic + pro) * 0.30  # 2.9% + $0.30/transaction
    net_rev = gross_rev - stripe_fee
    
    profit = net_rev - total_cost
    
    fill = WHITE if idx % 2 == 0 else LIGHT_BLUE
    
    values = [
        total,
        free, basic, pro,
        ai_cost, supa, vercel, other, total_cost,
        basic_rev, pro_rev, gross_rev, stripe_fee, net_rev,
        profit,
    ]
    
    for j, v in enumerate(values, 1):
        cell = ws2.cell(row=r, column=j, value=v)
        cell.fill = fill
        cell.font = B_REG
        cell.alignment = CENTER
        cell.border = THIN
        
        if j == 1:  # Total users
            cell.font = B_BOLD_BIG
            cell.alignment = CENTER
        elif j in (9,):  # Total cost
            cell.font = RED_BOLD
            cell.number_format = '$#,##0'
        elif j in (14,):  # Net revenue
            cell.font = GREEN_BOLD
            cell.number_format = '$#,##0'
        elif j == 15:  # Profit
            cell.font = GREEN_BOLD if v >= 0 else RED_BOLD
            cell.fill = PROFIT_GREEN if v >= 0 else LOSS_RED
            cell.number_format = '$#,##0'
        elif j in (5,6,7,8,10,11,12,13):
            cell.number_format = '$#,##0'

# Summary row
r += 1
style_row(ws2, r, 15, fill=NAVY)

# Margin calculation note
r += 2
ws2.merge_cells(f'A{r}:O{r}')
c = ws2.cell(row=r, column=1, value="📌 Gross Margin = (Net Revenue - AI Cost) / Net Revenue  |  This excludes marketing, salaries, and one-time setup costs")
c.font = Font(name='Calibri', size=10, color='666666')
c.alignment = CENTER

# ══════════════════════════════════════════════════════════════════════════
# SHEET 3: What If Conversion Is Different?
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Conversion Sensitivity")
ws3.sheet_properties.tabColor = 'FF6F00'

for i, w in enumerate([22, 12, 12, 12, 14, 14, 14, 14], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

r = 1
ws3.merge_cells('A1:H1')
c = ws3.cell(row=1, column=1, value="What If Conversion Rates Change?")
c.fill = NAVY; c.font = W_TITLE; c.alignment = CENTER
r = 2
ws3.merge_cells('A2:H2')
c = ws3.cell(row=2, column=1, value="Monthly P&L at 500 users with different free/basic/pro splits")
c.fill = NAVY; c.font = W_SUB; c.alignment = CENTER

r = 4
headers3 = ["Scenario", "Free %", "Basic %", "Pro %", "Total Cost", "Net Revenue", "PROFIT", "Margin"]
for j, h in enumerate(headers3, 1):
    cell = ws3.cell(row=r, column=j, value=h)
    cell.fill = DARK_HEADER; cell.font = W_BOLD; cell.alignment = CENTER

scenarios = [
    ("😰 Pessimistic", 85, 10, 5),
    ("🤔 Conservative", 75, 17, 8),
    ("📊 Base Case", 60, 25, 15),
    ("🚀 Optimistic", 50, 30, 20),
    ("🦄 Best Case", 40, 35, 25),
]

total_users = 500

for idx, (name, free_pct, basic_pct, pro_pct) in enumerate(scenarios):
    r += 1
    free = int(total_users * free_pct / 100)
    basic = int(total_users * basic_pct / 100)
    pro = total_users - free - basic
    
    ai = free * free_total + basic * basic_total + pro * pro_total
    infra = 25 + 20 + 21.25  # supa + vercel + other
    total_cost = ai + infra
    
    gross = basic * 9.99 + pro * 19.99
    stripe = gross * 0.029 + (basic + pro) * 0.30
    net = gross - stripe
    profit = net - total_cost
    margin = profit / net if net > 0 else 0
    
    fill = WHITE if idx % 2 == 0 else LIGHT_YELLOW
    if name == "📊 Base Case":
        fill = LIGHT_GREEN
    
    values = [name, f"{free_pct}%", f"{basic_pct}%", f"{pro_pct}%", total_cost, net, profit, margin]
    for j, v in enumerate(values, 1):
        cell = ws3.cell(row=r, column=j, value=v)
        cell.fill = fill
        cell.font = B_REG
        cell.alignment = CENTER
        if j in (5, 6):
            cell.number_format = '$#,##0'
        if j == 7:
            cell.font = GREEN_BOLD if isinstance(v, (int,float)) and v >= 0 else RED_BOLD
            cell.fill = PROFIT_GREEN if isinstance(v, (int,float)) and v >= 0 else LOSS_RED
            cell.number_format = '$#,##0'
        if j == 8 and isinstance(v, float):
            cell.number_format = '0%'

# ══════════════════════════════════════════════════════════════════════════
# SHEET 4: Cost Per User Breakdown (Simple)
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Unit Economics")
ws4.sheet_properties.tabColor = '1565C0'

for i, w in enumerate([24, 14, 14, 14, 14], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

r = 1
ws4.merge_cells('A1:E1')
c = ws4.cell(row=1, column=1, value="Unit Economics — Do We Make Money Per User?")
c.fill = NAVY; c.font = W_TITLE; c.alignment = CENTER

r = 3
headers4 = ["", "Free User", "Basic ($9.99)", "Pro ($19.99)", ""]
write_row(ws4, r, headers4, fill=DARK_HEADER, font=W_BOLD, align=CENTER)

data4 = [
    ("Monthly AI Cost", free_total, basic_total, pro_total),
    ("Monthly Subscription", 0, 9.99, 19.99),
    ("Stripe Fee (2.9%+$0.30)", 0, 9.99*0.029+0.30, 19.99*0.029+0.30),
    ("Net Revenue After Stripe", 0, 9.99-(9.99*0.029+0.30), 19.99-(19.99*0.029+0.30)),
    ("═══ PROFIT PER USER ═══", 0 - free_total, (9.99-(9.99*0.029+0.30)) - basic_total, (19.99-(19.99*0.029+0.30)) - pro_total),
]

for idx, (label, free_v, basic_v, pro_v) in enumerate(data4):
    r += 1
    is_total = "PROFIT" in label
    fill = NAVY if is_total else (WHITE if idx % 2 == 0 else LIGHT_BLUE)
    font = W_BOLD if is_total else B_REG
    
    for j, v in enumerate([label, free_v, basic_v, pro_v], 1):
        cell = ws4.cell(row=r, column=j, value=v)
        cell.fill = fill
        cell.font = font
        cell.alignment = LEFT if j == 1 else CENTER
        if j > 1:
            cell.number_format = '$#,##0.00'
            if is_total:
                cell.font = Font(name='Calibri', bold=True, size=12, 
                               color='4CAF50' if isinstance(v, (int,float)) and v >= 0 else 'FF5252')

# Bottom insight
r += 2
ws4.merge_cells(f'A{r}:E{r}')
net_basic = 9.99-(9.99*0.029+0.30) - basic_total
net_pro = 19.99-(19.99*0.029+0.30) - pro_total
c = ws4.cell(row=r, column=1, 
    value=f"💡 Each FREE user costs you ${free_total:.2f}/mo | Each BASIC user earns ${net_basic:.2f}/mo | Each PRO user earns ${net_pro:.2f}/mo")
c.font = Font(name='Calibri', bold=True, size=11, color='1B2A4A')
c.fill = LIGHT_GREEN; c.alignment = CENTER

r += 1
ws4.merge_cells(f'A{r}:E{r}')
# How many paying users needed to cover 1 free user?
paying_avg = (net_basic + net_pro) / 2
users_to_cover = free_total / paying_avg if paying_avg > 0 else 999
c = ws4.cell(row=r, column=1,
    value=f"💡 You need ~{users_to_cover:.1f} paying users to cover the cost of 1 free user")
c.font = Font(name='Calibri', size=10, color='666666')
c.alignment = CENTER


# ══════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════
out = "/Users/curvalux/NoteSnap/docs/budget/NoteSnap_Clean_PnL.xlsx"
wb.save(out)
print(f"Saved to {out}")
